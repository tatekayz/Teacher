# enroll by selecting names rather than right in then
# create login
# create an exe file
# create a database
# add and remove subjects
# generate leaners using class schedule
# edit and update all documents
# use inheritance
# use methods instead of functions

import sys
import pickle
import tkinter as tk
import random
import openpyxl as xl
import os
import datetime as d
from openpyxl.styles import Alignment, Border, Side, Font
import openpyxl.utils
from tkinter import filedialog


# tools
class Year:
    def __init__(self, year):
        self.year = year
        self.enrolled_pupils = {}
        self.enrollment = {}
        self.teachers = {}
        self.terms = {}
        self.stream_list = {'F_1': 1, 'F_2': 2, 'F_3': 4, 'F_4': 5, 'F_5': 5, 'F_6': 6}
        self.es: ''
        self.nos = ''
        self.transferred = {'in': {}, 'out': {}}


def prompt_list(prompt: str):
    prompt = prompt.split('\n')
    prompt = [x[2:].strip() for x in prompt]
    return prompt


def get_info(inf, y, s, t, w):
    inf.year = y.get()
    inf.stream = s.get()
    inf.term_part = t.get()
    w.destroy()


class DisApp:
    def __init__(self):
        self.txt = '98'
        self.opt = 99
        self.root = tk.Tk()
        self.root.geometry('280x580')
        self.root.resizable(True, True)
        self.log = ''

    def root_exist(self):
        try:
            self.root.winfo_exists()
            return True
        except tk.TclError:
            sys.exit(0)
            # return False


info = DisApp()


def destroy_widgets(root):
    for widget in root.winfo_children():
        widget.destroy()
        root.quit()


def get_t(root, num, opt_list):
    info.txt = opt_list[num]
    info.opt = 0
    destroy_widgets(root)


def get_data():
    while True:
        data = save_data({}, 'data_year', 1)
        year = get_int('year', 'year')
        info.log += '\n' + str(d.date.today()) + ' - ' + str(year)
        if len(str(year)) != 4 and year != 0:
            # info.log+= '\n' +str(d.time.now) +' - ' + str('GD0003 nyorai gore zvakanaka')
            continue
        if year not in data and year != 0:
            data[year] = Year(year)
        else:
            pass
            # info.log+= '\n' +str(d.time.now) +' - ' + str('year already in subjects')
        # create stream profile

        if year == 0 or year is None:
            return None, None
        else:
            return year, data


def get_tl(roots, num, opt_list, list_of_names):
    name = opt_list[num]
    root = roots[name]
    info.log += '\n' + str(d.date.today()) + ' - ' + str(len(list_of_names))
    info.log += '\n' + str(d.date.today()) + ' - ' + str(len(opt_list))
    info.log += '\n' + str(d.date.today()) + ' - ' + str(len(roots))
    if name in list_of_names:
        list_of_names.pop()
        root.config(fg="black")
    else:
        list_of_names.append(name)
        root.config(fg="blue")
    info.log += '\n' + str(d.date.today()) + ' - ' + str(len(list_of_names))
    info.log += '\n' + str(d.date.today()) + ' - ' + str(len(opt_list))
    info.log += '\n' + str(d.date.today()) + ' - ' + str(len(roots))


def get_op(root, num):
    info.opt = num + 1
    destroy_widgets(root)


def app_txt(opt_list):  # get_sub
    opt_list = [x for x in opt_list]
    buttons = []
    n = 0
    w = 40
    h = 40
    if 2 < len(opt_list) < 7:
        root = info.root
        if not info.root_exist():
            return info.txt
        for i in opt_list:
            button = tk.Button(root, text=i, command=lambda num=opt_list.index(i): get_t(root, num, opt_list),
                               width=round(w / 2), height=round(h / 3))
            buttons.append(button)
            button.grid(column=buttons.index(button) % 2, row=buttons.index(button) // 2)
            n += 1
        exit_button = tk.Button(root, text='Back', command=lambda: ext('opt', root), width=w)
        exit_button.grid(column=0, row=0, columnspan=2, rowspan=2)
        root.mainloop()
        return info.txt
    elif len(opt_list) < 3:
        root = info.root
        if not info.root_exist():
            return info.txt
        for i in opt_list:
            button = tk.Button(root, text=i, command=lambda num=opt_list.index(i): get_t(root, num, opt_list),
                               width=round(w / 2),
                               height=round(h / 3))
            buttons.append(button)
            button.grid(column=buttons.index(button) % 2, row=buttons.index(button) // 2)
            n += 1
        exit_button = tk.Button(root, text='Back', command=lambda: ext('opt', root), width=w)
        exit_button.grid(column=0, row=1, columnspan=2)
        root.mainloop()
        return info.txt
    elif len(opt_list) > 6:
        roots = {}
        root_num = -1
        for i in opt_list:
            root_num += 1
            count = root_num // 24
            if len(roots) == count:
                roots[count] = []
            roots[count].append(i)
        info.rt = 0
        info.opt = 200
        while True:
            rt = info.rt
            info.log += '\n' + str(d.date.today()) + ' - ' + str(rt)
            if rt in roots:
                root = info.root
                if not info.root_exist():
                    return info.txt
                r = 0
                c = 0
                for i in roots[rt]:
                    button = tk.Button(root, text=i, command=lambda num=opt_list.index(i): get_t(root, num, opt_list),
                                       width=round(w))
                    button.grid(row=r, column=c, columnspan=3)
                    r += 1
                next_page = tk.Button(root, text='next', command=lambda: next_p(rt, 0, roots, root),
                                      width=round(w / 3))
                prev_page = tk.Button(root, text='prev', command=lambda: next_p(rt, 1, roots, root),
                                      width=round(w / 3))
                ext_button = tk.Button(root, text='exit', command=lambda: ext('year', root), width=round(w / 3))
                next_page.grid(row=r, column=0)
                prev_page.grid(row=r, column=1)
                ext_button.grid(row=r, column=2)
                root.mainloop()
            opt = int(info.opt)
            if opt == 0:
                break
        return info.txt


def app_reg(opt_list1: list) -> list:
    opt_list = [x for x in opt_list1]
    del opt_list1
    list_of_names = []
    buttons = {}
    n = 0
    if opt_list is []:
        return list_of_names
    elif 2 < len(opt_list) < 7:
        root = info.root
        if not info.root_exist():
            return list_of_names
        for i in opt_list:
            button = tk.Button(root, text=i,
                               command=lambda num=opt_list.index(i): get_tl(buttons, num, opt_list, list_of_names),
                               width=13,
                               height=11)
            buttons[i] = button
            if i in list_of_names:
                buttons[i].config(fg='blue')
            button.grid(column=opt_list.index(i) % 2, row=opt_list.index(i) // 2)
            n += 1
        exit_button = tk.Button(root, text='submit', command=lambda: ext('opt', root), width=18)
        exit_button.grid(column=0, row=0, columnspan=2, rowspan=2)
        root.mainloop()
        del opt_list
        return list_of_names
    elif len(opt_list) < 3:
        root = info.root
        if not info.root_exist():
            return list_of_names
        for i in opt_list:
            button = tk.Button(root, text=i,
                               command=lambda num=opt_list.index(i): get_tl(buttons, num, opt_list, list_of_names),
                               width=13,
                               height=11)
            button.grid(column=opt_list.index(i) % 2, row=opt_list.index(i) // 2)
            buttons[i] = button
            if i in list_of_names:
                buttons[i].config(fg='blue')
            n += 1
        exit_button = tk.Button(root, text='submit', command=lambda: ext('opt', root), width=18)
        exit_button.grid(column=0, row=1, columnspan=2)
        root.mainloop()
        del opt_list
        return list_of_names
    elif len(opt_list) > 6:
        roots = {}
        root_num = -1
        for i in opt_list:
            root_num += 1
            count = root_num // 20
            if len(roots) == count:
                roots[count] = []
            roots[count].append(i)
        info.rt = 0
        info.opt = 200
        while True:
            rt = info.rt
            # info.log+= '\n' +str(d.time.now) +' - ' + str(rt)
            if rt in roots:
                root = info.root
                if not info.root_exist():
                    return list_of_names
                r = 0
                c = 0
                for i in roots[rt]:
                    button = tk.Button(root, text=i,
                                       command=lambda num=opt_list.index(i): get_tl(buttons, num, opt_list,
                                                                                    list_of_names),
                                       width=26, height=1)
                    button.grid(row=r, column=c)
                    buttons[i] = button
                    if i in list_of_names:
                        buttons[i].config(fg='blue')
                    r += 1
                next_page = tk.Button(root, text='next', command=lambda: next_p(rt, 0, roots, root),
                                      width=26)
                prev_page = tk.Button(root, text='prev', command=lambda: next_p(rt, 1, roots, root),
                                      width=26)
                ext_button = tk.Button(root, text='submit', command=lambda: ext('year', root), width=26)
                next_page.grid(row=r, column=0)
                prev_page.grid(row=r + 1, column=0)
                ext_button.grid(row=r + 2, column=0, columnspan=2)
                root.mainloop()
            opt = int(info.opt)
            if opt == 0:
                break
        del opt_list
        return list_of_names


def app_opt(opt_list, prompt):
    w = 40
    h = 40
    root = info.root
    if not info.root_exist():
        return info.opt
    tk.Label(root, text=prompt).grid(column=0, row=0, columnspan=2)
    buttons = []
    for i in opt_list:
        button = tk.Button(root, text=i, command=lambda num=opt_list.index(i): get_op(root, num), width=round(w / 2),
                           height=round(h / 3))
        buttons.append(button)
        button.grid(column=buttons.index(button) % 2, row=buttons.index(button) // 2 + 1)
    if len(opt_list) < 3:
        exit_button = tk.Button(root, text='Back1', command=lambda: ext('year', root), width=w)
        exit_button.grid(column=0, row=2, columnspan=2, rowspan=1)
    else:
        exit_button = tk.Button(root, text='Back2', command=lambda: ext('year', root), width=w)
        exit_button.grid(column=0, row=1, columnspan=2, rowspan=2)
    root.mainloop()
    return info.opt


def ins(root, entry, bv, l_txt):
    txt = entry.get()
    if not str.isdigit(txt):
        txt = ''
    entry.delete(0, tk.END)
    if bv in '0123456789':
        txt += bv
        entry.insert(0, txt)
    elif bv == 'clear':
        if len(txt) > 0:
            entry.insert(0, txt[:-1])
    elif bv == 'enter':
        if l_txt == 'year':
            year = str.strip(txt, '   --- /// ..')
            if str.isdigit(year) and len(year) == 4:
                year = int(year)
                info.txt = year
                destroy_widgets(root)
            else:
                entry.insert(0, 'error01 !!!! ' + year)
        else:
            if txt != '':
                info.txt = txt
                destroy_widgets(root)


def inp(root, entry, bv, l_txt):
    txt = entry.get()
    if txt == 'error01 !!!! ':
        txt = ''
    entry.delete(0, tk.END)
    if bv in '0123456789qwertyuiopasdfghjklzxcvbnm ':
        txt += bv
        entry.insert(0, txt)
    elif bv == 'clear':
        if len(txt) > 0:
            entry.insert(0, txt[:-1])
    elif bv == 'enter':
        if l_txt == 'year':
            year = str.strip(txt, '   --- /// ..')
            if str.isdigit(year) and len(year) == 4:
                year = int(year)
                info.txt = year
                destroy_widgets(root)
            else:
                entry.insert(0, 'error01 !!!! ')
        else:
            if txt != '':
                info.txt = txt
                destroy_widgets(root)


def ext(l_txt, root):
    if l_txt == 'mark':
        info.opt = 'e'
        info.txt = 'e'
        destroy_widgets(root)
    elif l_txt == 'year' or l_txt == 'name':
        info.opt = 0
        info.txt = 0
        destroy_widgets(root)
    elif l_txt == 'opt':
        info.opt = None
        info.txt = None
        destroy_widgets(root)


def skip(l_txt, root):
    if l_txt == 'mark':
        info.txt = 's'
        destroy_widgets(root)
    elif l_txt == 'year':
        info.txt = 0
        destroy_widgets(root)


def get_int(l_txt, prompt):
    root = info.root
    if not info.root_exist():
        return info.txt
    entry = tk.Entry(root)
    label1 = tk.Label(root, text=prompt)
    label2 = tk.Label(root, text=l_txt)
    entry.grid(row=1, column=1, columnspan=4)
    label1.grid(row=0, column=0, columnspan=4)
    label2.grid(row=1, column=0, )

    buttons = []
    n = 0
    w = 40
    h = 40
    opt_list = [x for x in '789456123']
    opt_list.extend(['enter', '0', 'clear'])
    for i in opt_list:
        button = tk.Button(root, text=i, command=lambda num=i: ins(root, entry, num, l_txt), width=round(w / 3),
                           height=round(h / 10))
        buttons.append(button)
        button.grid(column=buttons.index(button) % 3, row=(buttons.index(button) // 3) + 2)
        n += 1
    exit_button = tk.Button(root, text='exit', command=lambda: ext(l_txt, root), width=round(w * 2 / 3))
    skip_button = tk.Button(root, text='skip', command=lambda: skip(l_txt, root), width=round(w / 3))
    exit_button.grid(row=n, column=0, columnspan=2)
    skip_button.grid(row=n, column=2)
    root.mainloop()
    return info.txt


def app_input1(l_txt, prompt):
    root = info.root
    entry = tk.Entry(root, width=30)
    label1 = tk.Label(root, text=prompt)
    label2 = tk.Label(root, text=l_txt)
    entry.grid(row=1, column=2, columnspan=10)
    label1.grid(row=0, column=0, columnspan=10)
    label2.grid(row=1, column=0, columnspan=2)

    buttons = []
    n = 0
    w = 10
    h = 30
    opt_list = [x for x in '1234567890qwertyuiopasdfghjklzxcvbnm ']
    opt_list.insert(29, 'clear')
    opt_list.insert(39, 'enter')
    for i in opt_list:
        button = tk.Button(root, text=i, command=lambda num=i: inp(root, entry, num, l_txt), width=round(w / 10),
                           height=round(h / 30))
        buttons.append(button)
        if i == 'enter':
            button.config(width=4)
            button.grid(columnspan=3)
        button.grid(column=buttons.index(button) % 10, row=(buttons.index(button) // 10) + 2)
        n += 1
    exit_button = tk.Button(root, text='exit', command=lambda: ext(l_txt, root), width=18)
    exit_button.grid(row=n, column=2, columnspan=10)
    skip_button = tk.Button(root, text='skip', command=lambda: skip(l_txt, root), width=round(w / 10))
    skip_button.grid(row=n, column=0, columnspan=4)
    root.mainloop()
    del opt_list
    del buttons
    return info.txt


def cell_entry1(head, num=0):
    roots = {}
    root_num = -1
    dic = head.split('\n')
    if len(dic) > 0:
        for i in dic:
            root_num += 1
            count = root_num // 35
            if len(roots) == count:
                roots[count] = []
            roots[count].append(i)
        info.rt = 0
        while True:
            rt = info.rt
            # info.log+= '\n' +str(d.time.now) +' - ' + str(rt)
            if rt in roots:
                entries = {}
                root1 = info.root
                if not info.root_exist():
                    return
                row = num
                raw_line = roots[rt]
                if len(raw_line) > 0:
                    for item in raw_line:
                        row += 1
                        column = num
                        column_line = item.split('\t')
                        if len(column_line) > 0:
                            for i in column_line:
                                column += 1
                                if len(i.split(' ')) > 1 and len(i) > 3:
                                    e = tk.Entry(root1, width=10)
                                else:
                                    e = tk.Entry(root1, width=10)
                                entries[(row, column)] = e
                                e.insert(0, i)
                                e.grid(row=row, column=column)
                next_page = tk.Button(root1, text='next', command=lambda: next_p(rt, 0, roots, root1))
                prev_page = tk.Button(root1, text='prev', command=lambda: next_p(rt, 1, roots, root1))
                ext_button = tk.Button(root1, text='exit', command=lambda: ext('year', root1))
                next_page.grid(row=row + 1, column=1)
                prev_page.grid(row=row + 1, column=2)
                ext_button.grid(row=row + 1, column=3, columnspan=2)
                root1.mainloop()
            opt = int(info.opt)
            if opt == 0:
                break
    return


def app_display(dic):
    roots = {}
    root_num = -1
    if len(dic) > 0:
        for i in dic:
            root_num += 1
            count = root_num // 35
            if len(roots) == count:
                roots[count] = {}
            roots[count][i] = dic[i]
        info.rt = 0
        while True:

            rt = info.rt
            # info.log+= '\n' +str(d.time.now) +' - ' + str(rt)
            if rt in roots:
                root = info.root
                if not info.root_exist():
                    return
                r = 0
                c = 0
                for i in roots[rt]:
                    entry1 = tk.Entry(root)
                    entry1.insert(0, i)
                    entry2 = tk.Entry(root)
                    entry2.insert(0, roots[rt][i])
                    entry1.grid(row=r, column=c)
                    entry2.grid(row=r, column=c + 1)
                    r += 1
                next_page = tk.Button(root, text='next', command=lambda: next_p(rt, 0, roots, root))
                prev_page = tk.Button(root, text='prev', command=lambda: next_p(rt, 1, roots, root))
                ext_button = tk.Button(root, text='exit', command=lambda: ext('year', root))
                next_page.grid(row=r, column=0)
                prev_page.grid(row=r, column=1)
                ext_button.grid(row=r, column=0, columnspan=2)
                root.mainloop()
            opt = int(info.opt)
            if opt == 0:
                break
    return


def next_p(rt, num, roots, root):
    if num == 0:
        if -1 < rt < len(roots) - 1:
            # info.log+= '\n' +str(d.time.now) +' - ' + str(rt)
            rt += 1
            destroy_widgets(root)
            info.log += '\n' + str(d.date.today()) + ' - ' + str(f'hoyoo {rt}')
    elif num == 1:
        if 0 < rt < len(roots):
            destroy_widgets(root)
            rt -= 1
            info.log += '\n' + str(d.date.today()) + ' - ' + str('hoyoo2')
    info.rt = rt


def check_name(stud_entries, details, name):
    if len(details.get(name, '')) < 2 or 'enter' in details.get('name', ''):
        # details['name'] = ''
        stud_entries[name].delete(0, tk.END)
        stud_entries[name].insert(0, 'enter_proper_name')
        return 1
    else:
        full_name = ''
        for i in details['name'].split(' '):
            i = str.strip(i, '    .,  ;.     ')
            if str.isalpha(i):
                full_name += i.upper() + " "
        full_name = full_name[:-1]
        details['name'] = full_name
        stud_entries['name'].delete(0, tk.END)
        stud_entries['name'].insert(0, full_name)
        return 0


def check_details(stud_entries: dict, details: dict, value: int):
    """
    :param stud_entries:
    :param details:
    :param value:
    :return:
    """
    e = 0
    if value == 1:
        e += check_name(stud_entries, details, 'name')
        e += check_name(stud_entries, details, 'surname')
        if str.isdigit(details.get('level', '0')):
            if not 0 < int(details.get('level', '0')) < 7:
                stud_entries['level'].delete(0, tk.END)
                stud_entries['level'].insert(0, 'enter 1-6')
                e += 1
        else:
            e += 1
        dob = stud_entries['date_of_birth'].get()
        dob.strip()
        if str.isdigit(dob):
            if len(dob) != 8 or not 0 < int(dob[:2]) < 32 or not 0 < int(dob[2:4]) < 13:
                stud_entries['date_of_birth'].delete(0, tk.END)
                stud_entries['date_of_birth'].insert(0, 'enter in ddMMyyyy')
                e += 1
        else:
            stud_entries['date_of_birth'].delete(0, tk.END)
            stud_entries['date_of_birth'].insert(0, 'use digits')
            e += 1
        dob = stud_entries['date_of_admission'].get()
        dob.strip()
        if str.isdigit(dob):
            if len(dob) != 8 or not 0 < int(dob[:2]) < 31 or not 0 < int(dob[2:4]) < 13:
                stud_entries['date_of_admission'].delete(0, tk.END)
                stud_entries['date_of_admission'].insert(0, 'enter in ddMMyyyy')
                e += 1
        else:
            stud_entries['date_of_admission'] = d.date.today().strftime('%d%m%Y')
        sex = stud_entries['sex'].get()
        if str.isalpha(sex):
            if sex.lower() in 'm male boy men man':
                details['sex'] = 'M'
            elif sex.lower() in 'f female girl women lady woman':
                details['sex'] = 'F'
            else:
                stud_entries['sex'].delete(0, tk.END)
                stud_entries['sex'].insert(0, 'enter male or female')
                e += 1
        else:
            stud_entries['sex'].delete(0, tk.END)
            stud_entries['sex'].insert(0, 'enter male or female')
            e += 1
        return e
    else:
        return 0


def get_details(stud_entries, root, details, value):
    e = 0
    for item in stud_entries:
        # if stud_entries[item].get() != '':
        details[item] = stud_entries[item].get()
    # details['level'] = 1 #
    e += check_details(stud_entries, details, value)
    if e == 0:
        if details.get('name') is not None:
            details['name'] = str.strip(details['surname'].upper(), '   ,  _  ') + ' ' + str.strip(details['name'],
                                                                                                   '     ,.  b-  ')
            del details['surname']
            info.log += '\n' + str(d.date.today()) + ' - ' + str('its done')
        info.opt = 0
        destroy_widgets(root)


def collect_details(key: list, details: dict, value: int):
    """ creates an interface to collect data into a dictionary using the items in
    key as the keys of the dictionary
    :param key: str a list with the heading of the entries
    :param details: dict containing details of a pupil
    :param value: integer
    :return: dict updated
    """
    entry_names = key
    r = 0
    entry_columns = ['', '']
    stud_entries = {}
    root = info.root
    if not info.root_exist():
        return details
    # creates interface
    for i in enumerate(entry_columns):
        for j in enumerate(entry_names):
            r = j[0]
            c = i[0]
            if c == 0:
                entry = tk.Label(root, text=j[1])
            else:
                entry = tk.Entry(root)
                stud_entries[j[1]] = entry
                if details.get(j[1], -1) != -1:
                    entry.insert(0, details[j[1]])
                if j[1] == 'date_of_admission':
                    entry.insert(0,d.date.today().strftime('%d%m%Y'))
            entry.grid(row=r, column=c)
    submit = tk.Button(root, command=lambda: get_details(stud_entries, root, details, value), text='submit')
    exit_button = tk.Button(root, text='exit', command=lambda: ext('opt', root))
    exit_button.grid(row=r + 1, column=1)

    submit.grid(row=r + 1, column=0)
    root.mainloop()
    return details


def md(path, directory):
    path = os.path.dirname(os.path.dirname(path))
    path = os.path.join(path, directory)
    os.makedirs(path, exist_ok=True)
    return path


def save_data(data, data_name, opt):
    md(__file__, 'data')
    sys_path = os.path.dirname(os.path.dirname(__file__))
    path = os.path.join(sys_path, 'data', data_name)
    if opt == 0:
        pout = open(path, 'wb')
        nxt = data
        pickle.dump(nxt, pout)
        pout.close()
    elif opt == 1:
        if os.path.exists(path):
            pin = open(path, 'rb')
            date_year = pickle.load(pin)
            data = date_year
        else:
            info.log += '\n' + str(d.date.today()) + ' - ' + str('no file found')
            data = {}
        return data


def choose_book():
    root = info.root
    if not info.root_exist():
        return
    # root.withdraw()
    file_path = filedialog.askopenfilename(initialdir='C:\\Users\\Administrator\\Desktop\\system1\\app\\file',
                                           filetypes=[("xl files", "*.xlsx")])
    info.log += '\n' + str(d.date.today()) + ' - ' + str(file_path)
    destroy_widgets(root)
    return file_path
    pass


def cell_entry(wb, book_name, head, sheet, num):
    row = 0
    raw_line = head.split('\n')
    if len(raw_line) > 0:
        for item in raw_line:
            row += 1
            column = num
            column_line = item.split('\t')
            if len(column_line) > 0:
                for i in column_line:
                    column += 1
                    if str.isnumeric(i):
                        if "." in i:
                            sheet.cell(row, column).value = float(i)
                        else:
                            sheet.cell(row, column).value = int(i)
                    else:
                        sheet.cell(row, column).value = i
    wb.save(book_name)
    return sheet


def total_mark(overall):
    mark = 0
    for subject in overall:
        subject_mark = int(overall[subject])
        mark += subject_mark
        # info.log+= '\n' +str(d.time.now) +' - ' + str(f' mark is {mark}')
    return mark


def grade(num):
    grd = 2
    if 0 <= num < 50:
        grd = 0
    elif 50 <= num <= 100:
        grd = 1
    else:
        info.log += '\n' + str(d.date.today()) + ' - ' + str('G001 error')
    return grd


def sub_grade(overall):
    passed = 0
    for subject in overall:
        mark = int(overall[subject])
        passed += grade(mark)
    fail = len(overall) - passed
    return passed, fail


def avg(over, mark):
    if len(over) > 0:
        mark = mark / len(over)
        mark = round(mark, 2)
    else:
        mark = 0
    return mark


def ap2(key, details, value):
    w = 45
    opt_list = key
    roots = {}
    root_num = -1
    for i in opt_list:
        root_num += 1
        count = root_num // 28
        if len(roots) == count:
            roots[count] = []
        roots[count].append(i)
    info.rt = 0
    info.opt = 200
    while True:
        rt = info.rt
        # info.log+= '\n' +str(d.time.now) +' - ' + str(rt)
        if rt in roots:
            root = info.root
            if not info.root_exist():
                return details
            names = roots[rt]
            subs = ['', '']
            stud_entries = {}
            r = 0
            for i in enumerate(subs):
                for j in enumerate(names):
                    r = j[0]
                    c = i[0]
                    if c == 0:
                        entry = tk.Label(root, text=j[1])
                    else:
                        entry = tk.Entry(root)
                        stud_entries[j[1]] = entry
                        if details.get(j[1], -1) != -1:
                            entry.insert(0, details[j[1]])
                    entry.grid(row=r, column=c)
            next_page = tk.Button(root, text='next', command=lambda: next_p(rt, 0, roots, root),
                                  width=round(w / 3))
            submit = tk.Button(root, command=lambda: get_details(stud_entries, root, details, value),
                               text='submit', width=round(w / 3))

            prev_page = tk.Button(root, text='prev', command=lambda: next_p(rt, 1, roots, root),
                                  width=round(w / 3))
            ext_button = tk.Button(root, text='exit', command=lambda: ext('year', root), width=round(w / 3))
            next_page.grid(row=r + 1, column=0)
            prev_page.grid(row=r + 1, column=1)
            ext_button.grid(row=r + 2, column=1)
            submit.grid(row=r + 2, column=0)
            root.mainloop()
        opt = int(info.opt)
        if opt == 0:
            break
    return details


def dsp_marks1(class_info, subjects):
    roots = {}
    roots0 = {}
    roots1 = {}
    root_num = -1
    dic = class_info.students_list
    if len(dic) > 0:
        for i in dic:
            root_num += 1
            count = root_num // 10
            if len(roots0) == count:
                roots0[count] = []
            roots0[count].append(i)
        root_num = -1
        for j in subjects:
            root_num += 1
            count = root_num // 8
            if len(roots1) == count:
                roots1[count] = []
            roots1[count].append(j)
        for i in roots0.values():
            for j in roots1.values():
                roots[len(roots)] = (i, j)
        info.rt = 0
        info.txt = 200
        while True:

            rt = info.rt
            # info.log+= '\n' +str(d.time.now) +' - ' + str(rt)
            if rt in roots:
                root = info.root
                if not info.root_exist():
                    return
                students = class_info.students_list
                r = 0
                info.log += '\n' + str(d.date.today()) + ' - ' + str(roots)
                for stud in roots[rt][0]:
                    name_entries = tk.Entry(root, width=8)
                    name_entries.insert(0, stud)
                    overall = students[stud].overall
                    c = 0
                    r += 1
                    name_entries.grid(column=0, row=r)
                    for sub in roots[rt][1]:
                        sub_name = tk.Entry(root, width=4)
                        sub_name.insert(0, sub)
                        sub_entry = tk.Entry(root, width=4)
                        sub_entry.insert(0, overall.get(sub, 'xx'))
                        c += 1
                        sub_name.grid(column=c, row=0)
                        sub_entry.grid(column=c, row=r)
                r += 1
                next_page = tk.Button(root, text='next', command=lambda: next_p(rt, 0, roots, root))
                prev_page = tk.Button(root, text='prev', command=lambda: next_p(rt, 1, roots, root))
                ext_button = tk.Button(root, text='exit', command=lambda: ext('year', root))
                next_page.grid(row=r, column=0)
                prev_page.grid(row=r, column=1)
                ext_button.grid(row=r, column=0, columnspan=2)
                root.mainloop()
            opt = int(info.txt)
            if opt == 0:
                break


def dsp_marks(class_info, subjects):
    t = '\t'
    heading = ''
    students = class_info.students_list
    for s in subjects:
        heading += s[:3] + t
    heading += 'name'
    info.log += '\n' + str(d.date.today()) + ' - ' + str(heading)
    for stud in students:
        mark = ''
        overall = students[stud].overall
        for sub in subjects:
            mark += str(overall.get(sub, 'xx')) + t
        mark += students[stud].name
        info.log += '\n' + str(d.date.today()) + ' - ' + str(mark)


def dsp_bottom(clss):
    head = "bottom ten \n"
    top_ten = clss.bottom_ten
    try:
        positions = clss.class_position
    except AttributeError:
        positions = clss.overall_position
    for stud in top_ten:
        head += str(top_ten[stud]) + '\t' + str(stud) + '\t' + str(positions.get(stud, 'xx')) + "\n"
    return head


def marks_book(head, book_name):
    wb = xl.Workbook()
    sheet = wb.create_sheet('Sheet1')
    wb.create_sheet('Sheet2')
    del wb['Sheet']
    sheet = cell_entry(wb, book_name, head, sheet, 0)
    wb.save(book_name)
    return sheet


# enrollment
def get_results(level, details=None, name='0'):
    if details is None:
        details = {}
    if 0 < level < 5:
        key1 = 'English Mathematics Shona/Ndebele agriculture social_science physical_education total_units'
    elif 4 < level < 7:
        key1 = ('English Heritage Shona Maths Combined_science Physics Chemistry Biology English_literature '
                'Physical_Education Geography Practical Accounts Total_passes')
    else:
        key1 = ''
    key1 = key1.split(' ')
    if name == '0':
        results = collect_details(key1, details, 0)
    else:
        results = {x: '' for x in key1}
    return results


def gen_reg_no(enrollment, year):
    num = len(enrollment)
    letter = random.choice('abcdefghijklmnopqrstvwxyz')
    reg_num = 'R' + str(year)[-2:] + '0' * (4 - len(str(num))) + str(num) + letter.upper()
    return reg_num


class StreamProfile:
    def __init__(self, name, time, noc):
        self.name = name
        self.time = time
        if name[-1] in '1234':
            self.subjects = ['English', 'heritage', 'Family_and_religious_studies', 'Geography', 'Shona', 'Maths',
                             'Combined_science', 'biology', 'History', 'physics', 'Physical_education/lit', 'chemistry',
                             'Practicals', 'Accounts', 'business_studies']
            if name[-1] in '12':
                self.subjects = ['English', 'heritage', 'Geography', 'Shona', 'Maths', 'Combined_science',
                                 'History', 'Physical_education/lit', 'Practicals', 'Accounts', ]
                self.expected_pass = 6
            elif name[-1] in '34':
                self.expected_pass = 5
            self.classes = ['north', 'east', 'south', 'west']
        elif name[-1] in '56':
            self.subjects = ['pure_maths', 'statistics', 'physics', 'chemistry', 'biology', 'English',
                             'Shona', 'Geography', 'Practicals', 'heritage', 'Family_and_religious_studies', 'History',
                             'Accounts']
            self.expected_pass = 2
            self.classes = ['sciences', 'arts', 'commercials']
        self.noc = noc
        self.class_list = {}
        self.term_marks = {'1_mid': {}, '1_end': {}, '2_mid': {}, '2_end': {}, '3_mid': {}, '3_end': {}}


class ClassProfile:
    def __init__(self, name, year):
        self.name = name
        self.time = year
        self.registered_pupils = {}
        self.class_register = {}
        self.att_register = {'term_1': {}, 'term_2': {}, 'term_3': {}}
        self.term_marks = {'1_mid': {}, '1_end': {}, '2_mid': {}, '2_end': {}, '3_mid': {}, '3_end': {}}
        self.transferred = {'in': {}, 'out': {}}
        self.class_progress = ''


class RegisterProfile:
    def __init__(self, class_name):
        self.school_name = ''
        self.school_address = ''
        self.class_name = class_name
        self.period_from = ''
        self.period_to = ''
        self.register_number = ''
        self.Responsible_authority = ''
        self.registered_pupils = {}
        self.att_register = {'term_1': {}, 'term_2': {}, 'term_3': {}}


class StudentProfile:
    def __init__(self, student_name, class_number, class_name, stream_name, year, reg_number, enrollment):
        self.name = student_name
        self.year = year
        self.class_name = class_name
        self.stream_name = stream_name
        self.reg_number = reg_number
        self.class_number = class_number
        self.sex = enrollment[reg_number]['sex']
        self.dob = enrollment[reg_number]['date_of_birth']
        self.id_number = enrollment[reg_number]['id_number']
        self.doa = enrollment[reg_number]['date_of_admission']
        self.guardian_name = enrollment[reg_number]['guardian_name']
        self.phone_number = enrollment[reg_number]['phone_number']
        self.home_address = enrollment[reg_number]['home_address']
        self.reg_date = enrollment[reg_number]['date_of_admission']
        self.reg_results = enrollment[reg_number]['results']
        self.entry_type = enrollment[reg_number]['entry_type']
        self.religion = enrollment[reg_number]['religion']
        self.pic = ''
        self.attendance_totals = {}
        self.attendance_reg = {}
        self.subjects_registered = []
        self.bill = ''
        self.end_date = ''
        self.borrowed_books = []


def reg_opt(data, stream_name, class_name, year, enrollment):
    while True:
        register = data[year].stream_list[stream_name].class_list[class_name].class_register
        if register == {}:
            data[year].stream_list[stream_name].class_list[class_name].class_register = RegisterProfile(class_name)
        register = data[year].stream_list[stream_name].class_list[class_name].class_register.registered_pupils
        class_num_update(register, enrollment)
        opt = app_opt(prompt_list('1. create register \n'
                                  '2. create file \n'
                                  '3. edit register \n'
                                  '4. print register pupils\n'
                                  '5. transfer student'
                                  ), f'{class_name} {year} registration')
        if opt == 1:
            opt = app_opt(prompt_list('1. enter name \n'
                                      '2. enter reg_number\n'
                                      '3. load data \n'
                                      '4. select from enrolled \n5. change class'), f'{class_name} {year} registration')
            if opt == 1:
                create_register(data, stream_name, class_name, year, enrollment, r='name')
            elif opt == 2:
                create_register(data, stream_name, class_name, year, enrollment, r='reg_num')
            elif opt == 3:
                load_register(data, year, stream_name, class_name)
            elif opt == 4:
                create_register(data, stream_name, class_name, year, enrollment, r='list')



        elif opt == 2:
            create_file(data, year, stream_name, class_name)
        elif opt == 3:
            data = edit_register(data, year, stream_name, class_name)
        elif opt == 4:
            info.log += '\n' + str(d.date.today()) + ' - ' + str(class_name)
            register = data[year].stream_list[stream_name].class_list[class_name].class_register
            if register != {}:
                register = register.registered_pupils
            else:
                data[year].stream_list[stream_name].class_list[class_name] = RegisterProfile(class_name)
                register = register.registered_pupils
            app_display(register)

        elif opt == 5:
            register = data[year].stream_list[stream_name].class_list[class_name].class_register.registered_pupils
            stream = data[year].stream_list[stream_name]
            opt = app_opt(prompt_list('1. class transfer\n'
                                      '2. school transfer\n'
                                      '3. remove '), 'select transfer type')
            mode = opt - 1

            if mode == 0:
                student_name_list = app_reg(
                    [enrollment[x]['name'] for x in enrollment if enrollment[x].get('class_name', '') == class_name])

                new_stream = app_txt(data[year].stream_list)
                if new_stream in data[year].stream_list:
                    old_class = data[year].stream_list[stream_name].class_list[class_name]
                    class_list = data[year].stream_list[new_stream].class_list
                    new_class = app_txt(class_list)
                    if new_class in class_list:
                        for name in student_name_list:
                            reg_number = register[name].reg_number
                            stud_info = register.pop(name)
                            old_class.transferred['in'][name] = stud_info
                            enrollment[reg_number]['class_name'] = ''
                            new_register = data[year].stream_list[new_stream].class_list[
                                new_class].class_register.registered_pupils
                            name_entry(data[year].enrolled_pupils, new_register, name, reg_number, new_class,
                                       new_stream, year, enrollment)

            elif mode == 1:
                student_name_list = app_reg([enrollment[x]['name'] for x in enrollment])
                for name in student_name_list:
                    reg_number = data[year].enrolled_pupils.pop(name)
                    class_name = enrollment[reg_number]['class']
                    stream_name = 'F_' + class_name[0]
                    old_class = data[year].stream_list[stream_name].class_list[
                        class_name].class_register.registered_pupils
                    register = old_class.class_register.registered_pupils
                    stud_info = register.pop(name)
                    old_class.transferred['out'][name] = stud_info
                    data.tranferred['out'][reg_number] = enrollment.pop(reg_number)
            elif mode == 2:
                student_name_list = app_reg([enrollment[x]['name'] for x in enrollment])
                for name in student_name_list:
                    reg_number = data.enrolled_pupils.pop(name)
                    class_name = enrollment[reg_number].get('class', '')
                    if class_name != '':
                        stream_name = 'F_' + class_name[0]
                        old_class = data[year].stream_list[stream_name].class_list[
                            class_name].class_register.registered_pupils
                        register = old_class.class_register.registered_pupils
                        register.pop(name)
                    enrollment.pop(reg_number)
        else:
            break
    save_data(data, 'data_year', 0)
    return data


def class_num_update(reg, enrollment):
    boys = []
    girls = []
    for name in reg:
        reg[name].sex = enrollment[reg[name].reg_number]['sex']
        if reg[name].sex == 'M':
            boys.append(reg[name].name)
        elif reg[name].sex == 'F':
            girls.append(reg[name].name)
        else:
            info.log += '\n' + str(d.date.today()) + ' - ' + str('enter sex individual')
    order = sorted(girls)
    order.extend(sorted(boys))
    order1 = reg.copy()
    reg.clear()
    for name in order:
        reg[name] = order1.get(name, '')
        if reg.get(name) is not None:
            reg.get(name).class_number = order.index(name) + 1

    return


def name_entry(enroll, register, student_name, reg_number, class_name, stream_name, year, enrollment, f='0'):
    if student_name != '0' and reg_number != 0:
        if student_name not in enroll.values():
            if reg_number in enroll:
                student_name = enroll[reg_number]
            # reg_number = gen_reg_no(enroll,year)
            # enroll[reg_number] = student_name #id
            else:
                info.log += '\n' + str(d.date.today()) + ' - ' + str('student not enrolled go and register first')
        else:
            info.log += '\n' + str(d.date.today()) + ' - ' + str('name not found in enrolled pupils')

        if student_name in register.keys():  # or len(student_name)<4:
            info.log += '\n' + str(d.date.today()) + ' - ' + str(f'{student_name} already in register')

        else:
            if enrollment[reg_number].get('class_name', '') == '' or enrollment[reg_number].get('class_name',
                                                                                                '') == class_name:
                info.log += '\n' + str(d.date.today()) + ' - ' + str(
                    f'those in {class_name} {enrollment[reg_number].get('class_name', 'hapana')}')
                class_number = 0
                register[student_name] = StudentProfile(student_name, class_number, class_name, stream_name, year,
                                                        reg_number, enrollment)
                enrollment[reg_number]['class_name'] = class_name
                if f == 'xl':
                    sex_update(register, enrollment)
                class_num_update(register, enrollment)
                info.log += '\n' + str(d.date.today()) + ' - ' + str(
                    f'{student_name} is entered in re as {register[student_name].class_number}')
            else:
                info.log += '\n' + str(d.date.today()) + ' - ' + str(
                    f'class name error, {class_name}, {enrollment[reg_number].get('class_name', '')} ')
    else:
        info.log += '\n' + str(d.date.today()) + ' - ' + str('name or reg is 0')

    return register, student_name


def sex_update(register, enrollment):
    prev_name = '  '
    sex = 'F'
    change = 'M'
    for name in register.keys():
        if name[0] < prev_name[0]:
            sex, change = (change, sex)
        prev_name = name
        register[name].sex = sex
        reg_number = register[name].reg_number
        enrollment[reg_number]['sex'] = sex


# registration
def create_reg(enroll, register, class_name, stream_name, year, enrollment, r='name'):
    reg_number = 0
    while True:
        if r == 'name':
            student_name = app_input1('name', 'enter full name of student  \n')
            student_name = str(student_name).upper()
            if student_name == '0':
                break
            elif student_name in enroll:
                name_entry(enroll, register, student_name, reg_number, class_name, stream_name, year,
                           enrollment)
        elif r == "reg_num":
            reg_number = app_input1('reg_num', 'enter reg_number \n')
            reg_number = str.strip(reg_number.upper())
            student_name = enroll.get(reg_number, '')
            if reg_number in enrollment:
                name_entry(enroll, register, student_name, reg_number, class_name,
                           stream_name, year, enrollment)

        elif r == "list":
            student_name_list = app_reg(
                [enrollment[x]['name'] for x in enrollment if enrollment[x].get('class_name', '') == ''])
            for student_name in student_name_list:
                if student_name in enroll:
                    reg_number = enroll.get(student_name)
                    register, student_name = name_entry(enroll, register, student_name, reg_number,
                                                        class_name, stream_name, year, enrollment)
            break

    return register


def create_register(data, stream_name, class_name, year, enrollment, r="name"):
    register = data[year].stream_list[stream_name].class_list[class_name].class_register
    if register == {}:
        data[year].stream_list[stream_name].class_list[class_name].class_register = RegisterProfile(class_name)
    register = data[year].stream_list[stream_name].class_list[class_name].class_register.registered_pupils
    enroll = data[year].enrolled_pupils
    create_reg(enroll, register, class_name, stream_name, year, enrollment, r=r)
    save_data(data, 'data_year', 0)
    return data


def create_file(data, year, stream_name, class_name):
    register = data[year].stream_list[stream_name].class_list[class_name].registered_pupils
    book_name = class_name + str(year) + 'names.xlsx'
    sys_path = os.path.dirname(os.path.dirname(__file__))
    book_name = os.path.join(sys_path, 'files', book_name)
    ws = xl.Workbook()
    sheet = ws.create_sheet('Sheet1')
    del ws['Sheet']
    row = sheet.max_row
    for i in register:
        sheet.cell(row, 1).value = register[i].class_number
        sheet.cell(row, 2).value = i
        row += 1
    ws.save(str(book_name))
    info.log += '\n' + str(d.date.today()) + ' - ' + str('done')


def load_register(data, year, stream_name, class_name):
    while True:
        book_name = class_name + str(year) + 'names.xlsx'
        sys_path = os.path.dirname(os.path.dirname(__file__))
        book_name = os.path.join(sys_path, 'files', book_name)
        register = data[year].stream_list[stream_name].class_list[class_name].class_register.registered_pupils
        enroll = data[year].enrolled_pupils
        enrollment = data[year].enrollment
        if os.path.exists(book_name):
            wb = xl.load_workbook(str(book_name))
            sheet = wb['Sheet1']
            for i in range(1, sheet.max_row + 1):
                info.log += '\n' + str(d.date.today()) + ' - ' + str(sheet.max_row)
                num = sheet.cell(i, 1).value
                student_name = sheet.cell(i, 2).value
                if num is None:
                    num = '0'
                if student_name is None:
                    student_name = '0'
                # print (f' num is {num} , student_name')
                if str.isdecimal(str(num)) and not str.isdecimal(student_name):
                    pass
                elif str.isdecimal(student_name) and not str.isdecimal(num):
                    info.log += '\n' + str(d.date.today()) + ' - ' + str('in 2')
                    student_name, num = (num, student_name)
                else:
                    info.log += '\n' + str(d.date.today()) + ' - ' + str('correct your document')
                if student_name not in enroll.values():
                    continue
                register = name_entry(enroll, register, student_name, num, class_name, stream_name, year, enrollment)[0]
        else:
            info.log += '\n' + str(d.date.today()) + ' - ' + str('file not found !!!!')

        break
    sex_update(register, enrollment)
    return data


def edit_register(data, year, stream_name, class_name):
    while True:
        register = data[year].stream_list[stream_name].class_list[class_name].class_register.registered_pupils
        stud_name = app_txt(register)

        if stud_name not in register:
            break
        opt = app_opt(prompt_list(f'1. Edit name \n2. change class number \n3. delete student'), 'enter option')
        if opt == 1:
            name = app_txt(register)
            new_name = app_input1('name', 'Enter new name- ')
            if name != new_name:
                register[name] = new_name
            info.log += '\n' + str(d.date.today()) + ' - ' + str('name updated')

        elif opt == 2:
            for i in register:
                if register[i] == stud_name:
                    del register[i]
                    i = app_input1('name', 'Enter new class name- ')
                    register[i] = stud_name
                    info.log += '\n' + str(d.date.today()) + ' - ' + str('class name updated')
        elif opt == 3:
            for i in register:
                if register[i] == stud_name:
                    del register[i]
                    info.log += '\n' + str(d.date.today()) + ' - ' + str('name deleted')
        break
    return data


def get_sor(year, term, data, stream_name, class_name):  # get student_class overall, register
    class_list, students_class, overall, register = {}, {}, {}, {}
    c = 0
    try:
        stream = data[year].stream_list[stream_name].term_marks[term]
        class_list = stream.class_list
    except Exception as e:
        info.log += '\n' + str(d.date.today()) + ' - ' + str(e)
        c = 1
    if class_name is None or c == 1:
        info.log += '\n' + str(d.date.today()) + ' - ' + str('T001 error')

    else:
        clss = class_list[class_name]
        students_class = clss.students_list
        overall = clss.overall_marks
        register = data[year].stream_list[stream_name].class_list[class_name].class_register
        if register != {}:
            register = register.registered_pupils
        else:
            data[year].stream_list[stream_name].class_list[class_name].class_register = RegisterProfile(class_name)
            register = data[year].stream_list[stream_name].class_list[class_name].class_register.registered_pupils
    return students_class, overall, register


def ext_xl(subjects, data, stream_name, class_name, year, term):
    book_name = filedialog.askopenfilename(initialdir='C:\\Users\\Administrator\\Desktop\\system1\\app\\file',
                                           filetypes=[("xl files", "*.xlsx")])
    students_class, overall, register = get_sor(year, term, data, stream_name, class_name)
    book_name = str(book_name)
    if not os.path.exists(book_name):
        return students_class, overall

    try:
        wb = xl.load_workbook(book_name)
        info.log += '\n' + str(d.date.today()) + ' - ' + str('book loaded ')
    except Exception as e:
        info.log += '\n' + str(d.date.today()) + ' - ' + str(f'EX001 {e}')
        return students_class, register
    sheet = wb['Sheet1']
    list_nms = []  # list of names marks and sub
    stud = {}
    for co in range(2, sheet.max_column + 1):
        for ro in range(1, 5):
            v = sheet.cell(ro, co).value
            v = str(v)
            for sub in subjects:
                if v[:7].lower() == sub[:7].lower():
                    # info.log += '\n' +str(d.date.today()) +' - ' + str(v)
                    v = sub

            if v in subjects:
                sp = ro + 1
                # info.log += '\n' +str(d.date.today()) +' - ' + str(v)
                for r1 in range(sp, sheet.max_row + 1):
                    name = sheet.cell(r1, 1).value
                    if name is None:
                        continue
                    if 'average' not in name.split(' '):
                        # info.log += '\n' +str(d.date.today()) +' - ' + str(name)
                        stud[name] = {}

                    mark = sheet.cell(r1, co).value
                    sub = v
                    if mark is None or not str.isdigit(str(mark)) or 'average' in name.split(' '):
                        continue
                    item = [name, mark, sub]
                    list_nms.append(item)
                break
    for name in stud:
        level = int(class_name[0])
        enroll_pupils(data[year].enrollment, data[year].enrolled_pupils, year, name=name, level=level)
        data = create_register_xl(data, stream_name, class_name, year, name)
        for i in list_nms:
            stud[i[0]][i[2]] = i[1]
    return stud


def create_register_xl(data, stream_name, class_name, year, name):
    register = data[year].stream_list[stream_name].class_list[class_name].class_register
    info.log += '\n' + str(d.date.today()) + ' - ' + str(register)
    if register != {}:
        register = register.registered_pupils
    else:
        data[year].stream_list[stream_name].class_list[class_name].class_register = RegisterProfile(class_name)
        register = data[year].stream_list[stream_name].class_list[class_name].class_register.registered_pupils
    enroll = data[year].enrolled_pupils
    enrollment = data[year].enrollment
    reg_number = 'xx'
    while True:
        student_name = name
        if student_name not in enroll.values():
            break
        if student_name == '0':
            break
        for i in enroll:
            if student_name == enroll[i]:
                reg_number = i
        name_entry(enroll, register, student_name, reg_number, class_name, stream_name, year, enrollment, f='xl')
        break

    return data


def create_stream_profile(data, stream_name, year):
    name = stream_name
    noc = 4
    if data[year].stream_list[name] in [1, 2, 3, 4, 5, 6]:
        data[year].stream_list[name] = StreamProfile(name, year, noc)
        info.log += '\n' + str(d.date.today()) + ' - ' + str(f'{name} created\n')
    else:
        # info.log += '\n' +str(d.date.today()) +' - ' + str(f'{name} already recorded\n')
        pass
    return data


def create_class_profile(data, stream_name, name, year):
    if len(data[year].stream_list[stream_name].class_list) < data[year].stream_list[stream_name].noc:
        name = stream_name[-1] + '_' + name.upper()
        if stream_name not in data.get(year).stream_list.keys():
            info.log += '\n' + str(d.date.today()) + ' - ' + str(f'C002 error')
        else:
            data[year].stream_list[stream_name].class_list[name] = ClassProfile(name, year)
            # info.log += '\n' +str(d.date.today()) +' - ' + str(f'{name} created {stream_name}\n')
    else:
        info.log += '\n' + str(d.date.today()) + ' - ' + str('have reached max number of classes')
    return data


def get_stream(year, data, stream_name=''):
    while True:
        year_streams = data[year].stream_list
        info.log += '\n' + str(d.date.today()) + ' - ' + str(year) + ' streams'
        if stream_name == '':
            stream_name = app_txt([x for x in year_streams])
        if stream_name is None:
            return 0
        if stream_name == 0:
            return stream_name
        if stream_name == '0':
            return int(stream_name)

        if data[year].stream_list.get(stream_name) in [1, 2, 3, 4, 5, 6]:
            info.log += '\n' + str(d.date.today()) + ' - ' + str(f'creating new profile {stream_name}')
            data = create_stream_profile(data, stream_name, year)
            for name in data[year].stream_list[stream_name].classes:
                data = create_class_profile(data, stream_name, name, year)
        return stream_name


def enroll_data_ext(enrollment, enrolled_pupils, year, data):
    book_name = choose_book()
    if os.path.exists(book_name):
        wb = xl.load_workbook(book_name)
        test_key = ''
        sheet = wb['Sheet1']
        key = ('reg_num name id_number date_of_birth sex level entry_type date_of_admission guardian_name '
               'phone_number home_address religion results class_name')
        # print(sheet.max_column)
        if sheet.max_column < 13:
            return
        for i in range(1, sheet.max_column + 1):
            test_key += str(sheet.cell(1, i).value) + ' '
        if test_key[:30] == key[:30]:
            details = {}
            reg_num = ''
            for r in range(2, sheet.max_row):
                for c in range(1, sheet.max_column + 1):
                    v = sheet.cell(r, c).value
                    if v is None:
                        v = ""
                    if c == 1:
                        reg_num = v
                        details = {}
                        # info.log += '\n' +str(d.date.today()) +' - ' + str(v)
                        continue
                    else:
                        reg_num = ''

                    if sheet.cell(1, c).value is not None:
                        if sheet.cell(1, c).value == 'results':
                            details[sheet.cell(1, c).value] = eval(v)
                        else:
                            details[sheet.cell(1, c).value] = v
                if reg_num not in enrollment:
                    if reg_num == '':
                        reg_num = gen_reg_no(enrollment, year)
                    if details['name'] == '' or details['sex'] == '' or details['level'] == '':
                        continue
                    else:
                        enrollment[reg_num] = details
                        enrolled_pupils[reg_num] = enrollment[reg_num]['name']

                        if details.get('class_name', '') != '':
                            info.log += '\n' + str(d.date.today()) + ' - ' + str('class_name')
                            stream_name = 'F_' + details['class_name'][0]
                            stream_name = get_stream(year, data, stream_name=stream_name)
                            info.log += '\n' + str(d.date.today()) + ' - ' + str(stream_name)
                            class_name = details['class_name']
                            if data[year].stream_list[stream_name].class_list[class_name].class_register == {}:
                                data[year].stream_list[stream_name].class_list[
                                    class_name].class_register = RegisterProfile(class_name)
                            register = data[year].stream_list[stream_name].class_list[
                                class_name].class_register.registered_pupils
                            name_entry(enrolled_pupils, register, details['name'], reg_num, class_name, stream_name,
                                       year,
                                       enrollment)
                            # info.log += '\n' +str(d.date.today()) +' - ' + str(register)
                            pass
                else:
                    info.log += '\n' + str(d.date.today()) + ' - ' + str('already in ')
        else:
            # print('we are here')
            return
        return


def enroll_pupils(enrollment, enrolled_pupils, year, edit=False, name='0', sex='', level: int = ""):
    reg_num = ''
    if edit:
        name = app_txt(enrolled_pupils.values())
        if name in enrolled_pupils.values():
            for reg_num in enrolled_pupils:
                if enrolled_pupils[reg_num] == name:
                    break
            details = enrollment[reg_num]
            name = name.split(' ')
            details['surname'] = name[0]
            details['name'] = ''
            # details['level'] = 1
            name.pop(0)
            for i in name:
                details['name'] += i + ' '
            details['name'].rstrip(' ')
        else:
            return enrollment
    else:
        details = {}

    key = ('surname name id_number date_of_birth sex level entry_type date_of_admission guardian_name phone_number '
           'home_address religion')
    key = key.split(' ')
    if name == '0' and not edit:
        detail = collect_details(key, details, 1)

    elif edit:
        detail = collect_details(key, details, 1)

    else:
        if name not in enrolled_pupils.values():
            detail = {'name': name, 'id_number': '', 'date_of_birth': '', 'sex': sex, 'level': level, 'entry_type': '',
                      'date_of_admission': d.date.today().strftime('%d%m%Y'), 'guardian_name': '', 'phone_number': '',
                      'home_address': '', 'religion': ''}
        else:
            info.log += '\n' + str(d.date.today()) + ' - ' + str('kudhara akaisiwa munhu uyu')
            return enrollment, enrolled_pupils
    if len(detail) == 0 or len(detail['name']) < 2:
        return 0
    if edit:
        # info.log += '\n' +str(d.date.today()) +' - ' + str(detail)
        detail['results'] = get_results(int(detail['level']), details=detail['results'])
        enrollment[reg_num] = detail
        enrolled_pupils[reg_num] = enrollment[reg_num]['name']
    else:
        if detail['level'] == '':
            return 0
        detail['results'] = get_results(int(detail['level']), name=name)

    if len(detail['results']) != 0 and not edit:
        reg_num = gen_reg_no(enrollment, year)
        if reg_num not in enrollment:
            for i in enrollment.values():
                if detail['name'] == i['name']:
                    info.log += '\n' + str(d.date.today()) + ' - ' + str(f' {i["name"]} already registered in')
                    break
            if not edit:
                enrollment[reg_num] = detail
                enrolled_pupils[reg_num] = enrollment[reg_num]['name']
                info.log += '\n' + str(d.date.today()) + ' - ' + str('enrolled in enrollment is also true')

        else:
            info.log += '\n' + str(d.date.today()) + ' - ' + str('error reg number already in register')
            return enrollment
    return enrollment


def run1(year, data):
    while True:
        enrolled_pupils = data[year].enrolled_pupils
        enrollment = data[year].enrollment
        opt = app_opt(prompt_list('1. Enroll pupil\n'
                                  '2. class registration\n'
                                  '3. edit data\n'
                                  '4. print enrolled people\n'
                                  '5. remove student\n'
                                  '6. extract from file'), f'{year} Enrollment and Registration')
        if opt == 0:
            break
        if opt == 1:
            while True:
                r1 = enroll_pupils(enrollment, enrolled_pupils, year)
                if r1 == 0:
                    break
        elif opt == 2:
            while True:
                stream_name = get_stream(year, data)
                info.log += '\n' + str(d.date.today()) + ' - ' + str(f' aya  streams acho {stream_name}')
                if stream_name == 0 or stream_name is None:
                    break
                while True:
                    stream = data[year].stream_list.get(stream_name)
                    # info.log += '\n' +str(d.date.today()) +' - ' + str(data[year].stream_list)
                    class_name = app_txt(stream.class_list)
                    if class_name == 0 or class_name is None:
                        break
                    data = reg_opt(data, stream_name, class_name, year, enrollment)

        elif opt == 3:
            enroll_pupils(enrollment, enrolled_pupils, year, edit=True)
            pass
        elif opt == 4:
            info.log += '\n' + str(d.date.today()) + ' - ' + str('\nenrolled students')
            heading = 'reg_num'
            for reg_num in enrollment:
                for head in enrollment[reg_num]:
                    heading += '\t' + head
                break
            for reg_num in enrollment:
                name = str(reg_num)
                for head in enrollment[reg_num]:
                    name += '\t' + str(enrollment[reg_num][head])
                heading += '\n' + name
            book_name = str(year) + 'registration.xlsx'
            sys_path = os.path.dirname(os.path.dirname(__file__))
            md(__file__, (os.path.join(sys_path, 'files', str(year))))
            book_name = os.path.join(sys_path, 'files', str(year), book_name)
            cell_entry1(heading)
            wb = xl.Workbook()
            sheet = wb.create_sheet('Sheet1')
            del wb['Sheet']
            cell_entry(wb, book_name, heading, sheet, 0)
        elif opt == 5:
            student_name_list = app_reg([enrollment[x]['name'] for x in enrollment])
            for name in student_name_list:
                reg_number = ''
                class_name = ''
                if name in data[year].enrolled_pupils:
                    reg_number = data[year].enrolled_pupils.pop(name)
                    class_name = enrollment[reg_number].get('class', '')
                if class_name != '':
                    stream_name = 'F_' + class_name[0]
                    old_class = data[year].stream_list[stream_name].class_list[
                        class_name].class_register.registered_pupils
                    register = old_class.class_register.registered_pupils
                    register.pop(name)
                enrollment.pop(reg_number)
        elif opt == 6:
            enroll_data_ext(enrollment, enrolled_pupils, year, data)
        save_data(data, 'data_year', 0)


# mark register
class AttRegister:
    def __init__(self, class_name, term_name):
        self.class_name = class_name
        self.term_weeks_num = 13
        self.opening_day = ''
        self.closing_day = ''
        self.week_days = ''
        self.term_name = term_name
        self.date_of_first_friday = ''
        self.daily_register = {}
        self.weekly_totals = {}
        self.weekly_total = {}
        self.students_totals = {}
        self.ga_total = 0
        self.gp_total = 0

    def update_totals(self):
        self.weekly_totals.clear()
        self.weekly_total.clear()
        self.students_totals.clear()
        self.ga_total = 0
        self.gp_total = 0
        for name in self.daily_register:
            if name not in self.students_totals:
                self.students_totals[name] = {}
                self.students_totals[name]['a'] = 0
                self.students_totals[name]['p'] = 0
            if name not in self.weekly_total:
                self.weekly_total[name] = {}

            term_weeks = self.daily_register[name]
            for week in term_weeks:
                if week not in self.weekly_totals:
                    self.weekly_totals[week] = {}
                    self.weekly_totals[week]['a'] = 0
                    self.weekly_totals[week]['p'] = 0
                # attendance totals
                self.ga_total += [a for a in term_weeks[week].values()].count(0)
                self.gp_total += [b for b in term_weeks[week].values()].count(1)
                self.weekly_totals[week]['a'] += [c for c in term_weeks[week].values()].count(0)
                self.weekly_totals[week]['p'] += [c for c in term_weeks[week].values()].count(1)
                self.students_totals[name]['a'] += [ab for ab in term_weeks[week].values()].count(0)
                self.students_totals[name]['p'] += [pr for pr in term_weeks[week].values()].count(1)
                self.weekly_total[name][week] = [pr for pr in term_weeks[week].values()].count(1)


# run4 functions
def daily_reg_book_entry(term_reg, sheet, class_name, year):
    daily_register = term_reg.daily_register
    weekly_totals = term_reg.weekly_totals
    weekly_total = term_reg.weekly_total
    students_totals = term_reg.students_totals
    gp_total = term_reg.gp_total
    ga_total = term_reg.ga_total
    ro = 6
    week_num = 0
    if sheet.max_row == 1 and sheet.max_column == 1:
        sheet.cell(1, 1).value = class_name + str(year)
        sheet.cell(2, 1).value = 'week_number'
        sheet.cell(3, 1).value = 'week_ending'
        sheet.cell(4, 1).value = 'days'
        sheet.cell(5, 1).value = 'class_number'
        sheet.cell(5, 2).value = 'student_name'
        sheet.cell(ro + len(daily_register), 1).value = 'total attendance'
        sheet.cell(ro + len(daily_register) + 1, 1).value = 'total absence'
        sheet.cell(ro + len(daily_register) + 2, 1).value = 'total possible attendance'
        info.log += '\n' + str(d.date.today()) + ' - ' + str('done')
    co = 3
    for name in daily_register:
        term_weeks = daily_register[name]
        if sheet.cell(ro, 1).value is None:
            sheet.cell(ro, 1).value = ro - 5
            sheet.cell(ro, 2).value = name
            co = 3
            if sheet.max_column == 2:
                for week in term_weeks:
                    week_num += 1
                    sheet.cell(ro - 4, co).value = 'week_' + str(week_num)
                    sheet.merge_cells(start_row=ro - 4, end_row=ro - 4, start_column=co, end_column=co + 5)
                    sheet.cell(ro - 3, co).value = 'Fri_' + str(week)
                    sheet.merge_cells(start_row=ro - 3, end_row=ro - 3, start_column=co, end_column=co + 5)
                    for day in term_weeks[week]:
                        sheet.cell(ro - 2, co).value = day
                        sheet.column_dimensions[openpyxl.utils.get_column_letter(co)].width = 3
                        co += 1
                    sheet.cell(ro - 2, co).value = 'total'
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(co)].width = 5
                    co += 1
                sheet.cell(ro - 2, co).value = 'grand_total'
                sheet.column_dimensions[openpyxl.utils.get_column_letter(co)].width = 8
                co = 3

            for week in term_weeks:
                for day in term_weeks[week]:
                    sheet.cell(ro, co).value = term_weeks[week][day]
                    co += 1
                # info.log += '\n' +str(d.date.today()) +' - ' + str('weekly total',weekly_total)
                sheet.cell(ro, co).value = weekly_total[name][week]
                sheet.cell(6 + len(daily_register), co).value = weekly_totals[week]['p']
                sheet.cell(6 + len(daily_register) + 1, co).value = weekly_totals[week]['a']
                sheet.cell(6 + len(daily_register) + 2, co).value = weekly_totals[week]['a'] + weekly_totals[week]['p']
                co += 1
        sheet.cell(ro, co).value = students_totals[name]['p']
        ro += 1
    sheet.cell(ro, 1).value = 'total attendance'
    sheet.merge_cells(start_row=ro, end_row=ro, start_column=1, end_column=2)
    sheet.cell(ro, co).value = gp_total
    sheet.cell(ro + 1, 1).value = 'total absence'
    sheet.merge_cells(start_row=ro + 1, end_row=ro + 1, start_column=1, end_column=2)
    sheet.cell(ro + 1, co).value = ga_total
    sheet.cell(ro + 2, 1).value = 'total possible attendance'
    sheet.merge_cells(start_row=ro + 2, end_row=ro + 2, start_column=1, end_column=2)
    sheet.cell(ro + 2, co).value = ga_total + gp_total
    syd = Side(style="thin")
    b_s = Border(left=syd, right=syd, top=syd, bottom=syd)
    sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=sheet.max_column)
    sheet.merge_cells('A2:B2')
    sheet.merge_cells('A3:B3')
    sheet.merge_cells('A4:B4')
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['A'].width = 3
    for i in range(1, sheet.max_row + 1):
        sheet.row_dimensions[i].height = 10
        for j in range(1, sheet.max_column + 1):
            sheet.cell(i, j).border = b_s
            sheet.cell(i, j).font = Font(size=8,name='Times New Roman')


def daily_prof_book_entry(registered_pupils, sheet, class_name, year, enrollment):
    ro = 6
    if sheet.max_row == 1 and sheet.max_column == 1:
        sheet.cell(1, 1).value = class_name + str(year)
        sheet.merge_cells(start_row=1, end_row=4, start_column=1, end_column=2)
        sheet.cell(5, 1).value = 'class_number'
        sheet.cell(5, 2).value = 'student_name'
        info.log += '\n' + str(d.date.today()) + ' - ' + str('done')
    key = 'id_number date_of_birth sex entry_type date_of_admission guardian_name phone_number home_address religion'
    key = key.split(' ')

    for name in registered_pupils:
        if sheet.cell(ro, 1).value is None:
            sheet.cell(ro, 1).value = ro - 5
            sheet.cell(ro, 2).value = name
            co = 3
            if sheet.max_column == 2:
                for title in key:
                    sheet.cell(ro - 4, co).value = title
                    co += 1
                co = 3

            for title in key:
                sheet.cell(ro, co).value = enrollment[registered_pupils[name].reg_number][title]
                co += 1
        ro += 1
    syd = Side(style="thin")
    b_s = Border(left=syd, right=syd, top=syd, bottom=syd)
    sheet.column_dimensions['B'].width = 22
    sheet.column_dimensions['A'].width = 3
    sheet.column_dimensions['C'].width = 8
    sheet.column_dimensions['D'].width = 8
    sheet.column_dimensions['E'].width = 3
    sheet.column_dimensions['F'].width = 3
    sheet.column_dimensions['G'].width = 8
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 10
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 10
    for i in range(1, sheet.max_row + 1):
        sheet.row_dimensions[i].height = 10
        for j in range(1, sheet.max_column + 1):
            sheet.cell(i, j).border = b_s
            sheet.cell(i, j).font = Font(size=8,name='Times New Roman')


def date_format(date_str: str, date_opt: int):
    """
    opt 1 converts yyyymmdd to ddmmyyyy
    opt 2 converts ddmmyyyy to yyyymmdd
    :param date_str:
    :param date_opt:
    :return:
    """
    f = ''
    if date_opt == 1:
        f = date_str[4:] + date_str[2:4] + date_str[0:2]
    if date_opt == 0:
        f = date_str[6:] + date_str[4:6] + date_str[0:4]
    return f


def tick(daily_register1, name1, date_of_friday1, day, opt, root):
    if opt == 1:
        daily_register1[name1][date_of_friday1][day] = 1
        # root.destroy()
    elif opt == 2:
        daily_register1[name1][date_of_friday1][day] = 0
        # root.destroy()
    elif opt == 'e':
        info.opt = 0
    elif opt == 'h':
        for name1 in daily_register1:
            daily_register1[name1][date_of_friday1][day] = 'H'
            # root.destroy()
        info.opt = 0
    destroy_widgets(root)


def att_entry(class_list, term_reg, class_name, enrollment, year, data):
    info.log += '\n' + str(d.date.today()) + ' - ' + str(data[year].stream_list)
    att_reg = data[year].stream_list['F_' + class_name[0]].class_list[class_name].class_register.att_register
    daily_register = term_reg.daily_register
    term_weeks_num = 14
    # opening_day = ''
    # closing_day = ''
    name = ''
    week_days = {'Mon': '', 'Tue': '', 'Wed': '', 'Thu': '', 'Fri': ''}
    # info.log += '\n' +str(d.date.today()) +' - ' + str('this is daily reg', daily_register)
    if daily_register == {}:
        week_days = {'Mon': '', 'Tue': '', 'Wed': '', 'Thu': '', 'Fri': ''}
        while True:
            first_friday = get_int('date', 'enter the day  of friday of the term "ddMM"\n')
            if first_friday is not None:
                if len(str(first_friday)) == 4:
                    first_friday += str(year)
                    date_of_fri = d.date.fromisoformat(date_format(first_friday, 1))
                    info.log += '\n' + str(d.date.today()) + ' - ' + str(date_of_fri.ctime()[0:3])
                    pass
                else:
                    continue
            else:
                info.log += '\n' + str(d.date.today()) + ' - ' + str('incorrect date')
                continue
            # opening_day = input('enter the opening day term "ddMM"'
            # closing_day = input('enter the opening day term "ddMM"'
            if date_of_fri.ctime()[0:3] == "Fri":
                for name in class_list:
                    if name not in daily_register:
                        daily_register[name] = {}
                        date_of_fridays = d.date.fromisoformat(date_format(first_friday, 1))
                        for i in range(0, term_weeks_num):
                            daily_register[name][date_of_fridays.strftime('%d%m%Y')] = {day: '' for day in week_days}
                            date_of_fridays += d.timedelta(weeks=1)
                # for i in daily_register:
                # info.log += '\n' +str(d.date.today()) +' - ' + str(i, daily_register[i])
                break
            else:
                info.log += '\n' + str(d.date.today()) + ' - ' + str('enter the correct date of friday')
                continue

    else:
        name = [x for x in daily_register][0]
        if name == '':
            return
        first_friday = [x for x in daily_register[name]][0]
    # reg marking

    while True and name != '':
        date_of_friday = app_txt(daily_register[name])
        if len(str(date_of_friday)) == 8:
            pass
        elif date_of_friday == '0' or date_of_friday == 'e':
            break
        else:
            continue

        day = app_txt(week_days)
        if day == '0' or day == 'e' or day is None:
            break
        info.opt = 200
        for name in class_list:
            if daily_register.get(name, 'not') == 'not':
                daily_register[name] = {}
                date_of_fridays = d.date.fromisoformat(date_format(first_friday, 1))
                for i in range(0, term_weeks_num):
                    daily_register[name][date_of_fridays.strftime('%d%m%Y')] = {day: '' for day in week_days}
                    date_of_fridays += d.timedelta(weeks=1)
                info.log += '\n' + str(d.date.today()) + ' - ' + str('name not')
            # opt = int(input(f'{day}\n {name} \n 1.present \n 2.absent \n 3.special\n'))
            if daily_register[name][date_of_friday][day] == '':
                root = info.root
                w = 40
                h = 40
                ex_b = tk.Button(root, text='exit',
                                 command=lambda: tick(daily_register, name, date_of_friday, day, 'e', root), width=w)
                present = tk.Button(root, text='present',
                                    command=lambda: tick(daily_register, name, date_of_friday, day, 1, root),
                                    width=round(w / 2), height=round(h * 2 / 3))
                absent = tk.Button(root, text='absent',
                                   command=lambda: tick(daily_register, name, date_of_friday, day, 2, root),
                                   width=round(w / 2), height=round(h * 2 / 3))
                holiday = tk.Button(root, text='holiday',
                                    command=lambda: tick(daily_register, name, date_of_friday, day, 'h', root),
                                    width=w)
                label = tk.Label(root,
                                 text=f'{day} week ending {date_of_friday[:2]}-{date_of_friday[2:4]}-'
                                      f'{date_of_friday[4:]}\n {name}')
                label.grid(row=1, column=1, columnspan=3)
                present.grid(row=2, column=1)
                absent.grid(row=2, column=3)
                holiday.grid(row=3, column=1, columnspan=3)
                ex_b.grid(row=4, column=1, columnspan=3)
                root.mainloop()
                if info.opt == 0:
                    break

        book_name = class_name + str(year) + 'attendance_record_book' + '.xlsx'
        md(__file__, 'file')
        sys_path = os.path.dirname(os.path.dirname(__file__))
        path = os.path.join(sys_path, 'file')
        book_name = os.path.join(path, book_name)
        wb = xl.Workbook()
        student_profile_sheet = wb['Sheet']
        attendance_sheets = {'term_1': wb.create_sheet('Sheet1'), 'term_2': wb.create_sheet('Sheet2'),
                             'term_3': wb.create_sheet('Sheet3')}
        term_reg.update_totals()
        # del wb.sheets['Sheet']
        info.log += '\n' + str(d.date.today()) + ' - ' + str(term_reg.weekly_totals)
        daily_prof_book_entry(class_list, student_profile_sheet, class_name, year, enrollment)
        for term in att_reg:
            if att_reg[term] != {}:
                daily_reg_book_entry(att_reg[term], attendance_sheets[term], class_name, year)

        save_data(data, 'data_year', 0)
        wb.save(book_name)
        break


def run4(year, stream_name, term, class_name, data):
    info.log += '\n' + str(d.date.today()) + ' - ' + str(term)
    class_list = data[year].stream_list[stream_name].class_list
    if class_list[class_name].class_register != {}:
        att_reg = class_list[class_name].class_register.att_register
    else:
        return
    term_reg = att_reg['term_' + term[0]]
    info.log += '\n' + str(d.date.today()) + ' - ' + str('this is term reg' + str(term_reg))
    if term_reg == {}:
        class_list[class_name].class_register.att_register['term_' + term[0]] = AttRegister(class_name, term)
        term_reg = class_list[class_name].class_register.att_register['term_' + term[0]]
        info.log += '\n' + str(d.date.today()) + ' - ' + str('ehe')
    registered_pupils = class_list[class_name].class_register.registered_pupils
    enrollment = data[year].enrollment
    opt = app_opt(prompt_list('1. mark register \n2.edit profile'), 'hoyoo')
    if opt == 1:
        class_info = class_list[class_name]
        # daily_register = class_info.att_register[term]
        att_entry(class_info.class_register.registered_pupils, term_reg, class_name, enrollment, year, data)

    if opt == 2:
        pass
        student_name = app_txt(registered_pupils)
        detail = enrollment[registered_pupils[student_name].reg_number]
        key = ('surname name id_number date_of_birth'
               ' sex level entry_type date_of_admission guardian_name phone_number home_address religion')
        key = key.split(' ')
        while True:
            detail = collect_details(key, detail, 1)
            if len(detail) == 0 or len(detail['name']) < 2:
                continue
            break


# term marks
class ClassTermMark:
    def __init__(self, name, year):
        self.name = name
        self.time = year
        self.part = ''
        self.term = ''
        self.overall_position = {}
        self.pass_number = ''
        self.fail_number = ''
        self.subjects = {}
        self.overall_marks = {}
        self.average_marks = {}
        self.class_position = {}
        self.pass_rate = 0
        self.sub_av = {}
        self.students_list = {}
        self.top_ten = ''
        self.bottom_ten = ''
        self.compiled = False


class StreamTermMark(ClassTermMark):  # number_of_classes
    def __init__(self, name, year):
        super().__init__(name, year)
        self.class_list = {}


def class_update(student_object):
    overall = student_object.overall
    passed, failed = sub_grade(overall)
    overall_mark = total_mark(overall)
    average_mark = avg(overall, overall_mark)
    student_object.passed = passed
    student_object.failed = failed
    student_object.overall_mark = overall_mark
    student_object.average_mark = average_mark


class StudentTermMark:
    def __init__(self, number, name):  # , cm):
        self.name = name
        self.class_number = number
        self.overall = {}
        self.class_position = 0
        self.overall_position = 0
        self.overall_mark = 0
        self.passed = 0
        self.average_mark = 0
        self.failed = 0


def record_mark_app(students_class, register, subjects, name, sub, mark2):
    # info.log += '\n' +str(d.date.today()) +' - ' + str(students_class)
    if name not in students_class.keys():
        info.log += '\n' + str(d.date.today()) + ' - ' + str(name)
        students_class[name] = StudentTermMark(register[name].class_number, name)
    mark1 = students_class[name].overall.get(sub, -1)
    if mark1 > 0:
        pass
    else:
        while True:
            mark1 = mark2
            if str.isdecimal(mark1):
                mark = int(mark1)
                if mark > 100:
                    info.log += '\n' + str(d.date.today()) + ' - ' + str('error value surpasses the overall')
                    break  #
                break
            else:
                mark = -1
                info.log += '\n' + str(d.date.today()) + ' - ' + str('\nThere was an error retry\n')
                if mark1 == 'e':
                    break
                elif mark1 == 's':
                    break
                break
        if sub in subjects and 0 <= mark <= 100:
            students_class[name].overall[sub] = mark
            class_update(students_class[name])
            info.log += '\n' + str(d.date.today()) + ' - ' + str('done!!!')
    return students_class[name]


def record_marks(students_class, register, subjects, name, sub):  # num is now name
    if name not in students_class.keys():
        students_class[name] = StudentTermMark(register[name], name)
    mark1 = students_class[name].overall.get(sub, -1)

    if mark1 > 0:
        pass
    else:
        while True:
            mark1 = get_int('mark', f'{sub} \n {name} ')
            if str.isdecimal(mark1):
                mark = int(mark1)
                if mark > 100:
                    info.log += '\n' + str(d.date.today()) + ' - ' + str('error value surpasses the overall')
                    continue
                break
            else:
                mark = -1
                info.log += '\n' + str(d.date.today()) + ' - ' + str('\nM0001 error\n')
                if mark1 == 'e':
                    return 0
                elif mark1 == 's':
                    break
                continue
        # info.log += '\n' +str(d.date.today()) +' - ' + str(num)
        if sub in subjects and 0 <= mark <= 100:
            students_class[name].overall[sub] = mark
            class_update(students_class[name])
            # info.log += '\n' +str(d.date.today()) +' - ' + str(students_class[num])
            info.log += '\n' + str(d.date.today()) + ' - ' + str('done!!!')
    return students_class[name]


def enter_marks(students_class, register, subjects, overall):
    info.log += '\n' + str(d.date.today()) + ' - ' + str(overall)
    # entering all marks for all students
    brk = 0
    for name in register:
        # info.log += '\n' +str(d.date.today()) +' - ' + str(students_class)
        if brk == 1:
            break
        for sub in subjects:
            student = record_marks(students_class, register, subjects, name, sub)
            if student == 0:
                brk = 1
                break
            overall[student.name] = student.overall
    info.log += '\n' + str(d.date.today()) + ' - ' + str('all done!!\n')
    return


def sub_mark(students_class, register, subjects, overall):
    num = app_txt(register)
    if num is None or num not in register:
        return students_class, overall
    for sub in subjects:
        student = record_marks(students_class, register, subjects, num, sub)
        if student == 0:
            break
        overall[student.class_number] = student.overall
    return students_class, overall


def sub_marks(students_class, register, subjects, overall):
    sub = app_txt(subjects)
    if sub in ['exit', 'e', 0, '0']:
        return students_class, overall
    dts = {name: overall[name].get(sub, '') for name in overall if name in register.keys()}
    mark_for_studs = ap2([x for x in register], dts, 0)
    for name in mark_for_studs:
        student = record_mark_app(students_class, register, subjects, name, sub, mark_for_studs[name])
        if student == 0:
            break
        overall[student.name] = student.overall
    return


def edit_mark(students_class, subjects, overall):
    while True:
        name = app_txt(students_class)
        if name == 0:
            return students_class, overall
        if name == 'not in register':
            info.log += '\n' + str(d.date.today()) + ' - ' + str(name)
            continue
        while True:
            sub = app_txt(subjects)
            if sub == 0 or sub is None:
                return students_class, overall
            mark = students_class[name].overall.get(sub, 'xx')
            mark = get_int('mark', f'current mark for {sub} is {mark} ')
            if mark == 'e':
                return students_class, overall
            else:
                try:
                    mark = int(mark)
                except ValueError:
                    info.log += '\n' + str(d.date.today()) + ' - ' + str('enter digits mhani iwe ah!')
                    continue
            if sub in subjects and 0 <= mark <= 100:
                students_class[name].overall[sub] = mark
                class_update(students_class[name])
            student = students_class[name]
            overall[name] = student.overall
            info.log += '\n' + str(d.date.today()) + ' - ' + str('all done \n')
            return


def create_ov(entries, students_class, register, subjects, overall, root):
    stud_overalls = {}
    for sub in entries:
        sub_marks_list = entries[sub]
        for stud in sub_marks_list:
            if stud not in stud_overalls:
                stud_overalls[stud] = {}
            mark = sub_marks_list[stud].get()
            if mark != '':
                stud_overalls[stud][sub] = sub_marks_list[stud].get()
    for i in stud_overalls:
        # info.log += '\n' +str(d.date.today()) +' - ' + str(i,stud_overalls[i])
        for j in stud_overalls[i]:
            mark = stud_overalls[i][j]
            sub = j
            students_class[i] = record_mark_app(students_class, register, subjects, i, sub, mark)
            overall[i] = students_class[i].overall

    root.destroy()
    return stud_overalls


def ap(student_class, register, subjects, overall):
    max_r = 25
    max_c = 7
    roots = []
    subs1 = []
    subs_all = []
    nams1 = []
    nams_all = []
    for h in range(0, (len(subjects) // max_c) + 1):
        subs = []
        for i in subjects:
            if i not in subs_all:
                subs.append(i)
                subs_all.append(i)
            if len(subs) == max_c:
                break
        subs1.append(subs)
    for j in range(0, (len(register) // max_r) + 1):
        nams = []
        for i in register:
            if i not in nams_all:
                nams.append(i)
                nams_all.append(i)
            if len(nams) == max_r:
                break
        nams1.append(nams)
    for i in subs1:
        for j in nams1:
            if (i, j) not in roots:
                roots.append((i, j))
    for thing in roots:
        names = ['']
        r = 0
        names.extend(thing[1])
        subs = ['']
        subs.extend(thing[0])
        stud_entries = {}
        root = tk.Tk()
        # creates interface
        for i in enumerate(subs):
            for j in enumerate(names):
                r = j[0]
                c = i[0]
                if c == 0:
                    entry = tk.Entry(root)
                    entry.insert(0, j[1])
                elif r == 0:
                    entry = tk.Entry(root)
                    entry.insert(0, i[1])
                    stud_entries[i[1]] = {}
                else:
                    entry = tk.Entry(root)
                    stud_entries[i[1]][j[1]] = entry
                entry.grid(row=r, column=c)
        for sub in stud_entries:
            for name in stud_entries[sub]:
                # info.log += '\n' +str(d.date.today()) +' - ' + str(f' entries{stud_entries} \n overall{overall}')
                if overall.get(name) is not None:
                    if overall[name].get(sub, '') != '':
                        mark = overall[name][sub]
                        stud_entries[sub][name].insert(0, mark)

        submit = tk.Button(root, command=lambda: create_ov(stud_entries, student_class, register, subs, overall, root),
                           text='submit')
        submit.grid(row=r + 1, column=0)
        root.mainloop()


def sub_marks3(students_class, register, subjects, overall):
    sub = app_txt(subjects)
    if sub in ['exit', 'e', 0, '0']:
        return students_class, overall
    for name in register:
        student = record_marks(students_class, register, subjects, name, sub)
        if student == 0:
            break
        overall[student.name] = student.overall
    return


def class_position(pupils):
    highest = -10000
    highest_name = ''
    position = {}
    r = len(pupils)
    for i in range(0, r):
        for name in pupils:
            if pupils[name] > highest:
                highest = pupils[name]
                highest_name = name
        position[highest_name] = highest
        pupils.pop(highest_name)
        highest = -10000  # set highest to lowest number
    # info.log += '\n' +str(d.date.today()) +' - ' + str(position)
    pos_num = {}
    num = 1
    score = {}
    for name in position:
        # info.log += '\n' +str(d.date.today()) +' - ' + str(len(pos_num))
        if len(pos_num) == 0:
            pos_num[name] = num
            score[str(position[name])] = num

        else:
            item = score.popitem()
            if item[0] == str(position[name]):
                pos_num[name] = item[1]
                score[item[0]] = item[1]

            else:
                pos_num[name] = num
                score[str(position[name])] = num
        num += 1
    return pos_num  # ,position,score


def compile_stud1(data, stream_name, subjects, term, expected_pass, year):
    stream = data[year].stream_list[stream_name].term_marks.get(term)
    ov_position_list = {}  # for student positions
    ov_position_list1 = {}  # for bottom ten creation
    ov_position_list2 = {}
    ov_pass_rate = 0.0
    sabas = {}
    for clss in stream.class_list.values():
        # CHECK IF ALL MARKS WERE ENTERED
        saba = {s: [] for s in subjects}
        register = data[year].stream_list[stream_name].class_list[clss.name].class_register
        if register != {}:
            register = register.registered_pupils
        else:
            continue
        students_list = clss.students_list
        error = 0
        error_names = []
        for student in students_list:
            # info.log += '\n' +str(d.date.today()) +' - ' + str(student)
            if len(students_list[student].overall) != len(subjects):
                error += 1  # 1  # student positioning
                error_names.append(f'Error 00: {students_list[student].name} has missing sub_marks')
        if len(students_list) != len(register):
            error += 1
            error_names.append("Error 01: some students are missing")

        if error > 0:
            if len(error_names) != 0:
                info.log += '\n' + str(d.date.today()) + ' - ' + str(f'missing marks for {error_names}\n\n')

        # PROCESSING OF DATA
        clss.class_position.clear()
        position_list = clss.class_position  # for student positions
        position_list1 = {}  # for bottom ten creation
        position_list2 = {}
        pass_rate = 0.0
        for person in students_list:  # to get two list used to create the bottom ten and top ten
            name = students_list[person].name
            mark = students_list[person].average_mark
            position_list[name] = mark
            ov_position_list[name] = mark
            clss.average_marks[name] = mark
            stream.average_marks[name] = mark
            if students_list[person].passed < expected_pass:
                position_list1[name] = 0 - mark  # list for creating bottom ten using negative value
                ov_position_list1[name] = 0 - mark
            else:
                pass_rate += 1
                ov_pass_rate += 1
                position_list2[name] = mark
                ov_position_list2[name] = mark
        clss.class_position = class_position(position_list.copy())  # creates a dictionary with position numbers
        top = class_position(position_list2)
        bottom = class_position(position_list1)  # create a dictionary with position starting with the least score
        top_ten = {}
        bottom_ten = {}
        for item in top:
            if top[item] not in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
                continue
            top_ten[item] = top[item]
        clss.top_ten = top_ten

        for item in bottom:
            bottom_ten[item] = bottom[item]
        clss.bottom_ten = bottom_ten

        for item in students_list:
            for sub in students_list[item].overall:
                saba.get(sub, []).append(students_list[item].overall.get(sub, 0))
        sabas[clss.name] = saba

        clss.pass_number = pass_rate
        clss.fail_number = len(clss.class_position) - pass_rate
        if len(students_list) > 0:
            pass_rate /= len(clss.class_position)
        else:
            pass_rate = 0.0
        pass_rate = round(pass_rate * 100,
                          2)  # info.log += '\n' +str(d.date.today()) +' - ' + str(f'"the pass rate is {pass_rate}"'
        clss.pass_rate = pass_rate
        info.log += '\n' + str(d.date.today()) + ' - ' + str(f'data compiled')
        clss.compiled = True
        # return clss
    stream.overall_position = class_position(ov_position_list.copy())  # creates a dictionary with position numbers
    top = class_position(ov_position_list2)
    bottom = class_position(ov_position_list1)  # create a dictionary with position starting with the least score

    top_ten = {}
    bottom_ten = {}
    for item in top:
        if top[item] > 10:
            break
        top_ten[item] = top[item]
    stream.top_ten = top_ten
    for item in bottom:
        bottom_ten[item] = bottom[item]
    stream.bottom_ten = bottom_ten

    avg1_mark = {}
    ll = {}
    for cm in sabas:
        saba = sabas[cm]
        clss = stream.class_list[cm]
        for s in saba:
            avg1_mark[s] = avg1_mark.get(s, 0)
            ll[s] = ll.get(s, 0)
            av_mark = 0.0
            for mark in saba[s]:
                av_mark += int(mark)
                avg1_mark[s] += int(mark)
                ll[s] += 1
            if len(saba[s]) != 0:
                av_mark /= len(saba[s])  # number of items in the mark list for each subject
            if ll[s] != 0:
                stream.sub_av[s] = round(avg1_mark[s] / ll[s])
            av_mark = round(av_mark)
            clss.sub_av[s] = av_mark

    stream.pass_number = ov_pass_rate
    stream.fail_number = len(stream.overall_position) - ov_pass_rate
    if len(stream.overall_position) > 0:
        ov_pass_rate /= len(stream.overall_position)
    else:
        ov_pass_rate = 0.0
    ov_pass_rate = round(ov_pass_rate * 100,
                         2)  # info.log += '\n' +str(d.date.today()) +' - ' + str(f'"the pass rate is {pass_rate}"'
    stream.pass_rate = ov_pass_rate

    for item in stream.class_list:
        cl = stream.class_list[item]
        students_list = cl.students_list
        # update position
        # info.log += '\n' +str(d.date.today()) +' - ' + str(pos_number[i], i)
        for person in students_list:  # update position
            students_list[person].position = cl.class_position.get(person, '')  # class positiom
            students_list[person].overall_position = stream.overall_position[person]  # stream positon
            cl.overall_position[person] = stream.overall_position[person]


def mark_opt(data, students_class, stream_name, class_name, year, term, register, subjects, overall,
             expected_pass):
    while True:
        if len(register) < 1:
            info.log += '\n' + str(d.date.today()) + ' - ' + str('no students registered')
            return data

        opt = app_opt(
            prompt_list(
                '1.Enter all marks for all students and all subjects \n2.Enter marks for student all subjects '
                '\n3.Edit marks for student per subject \n4.Enter marks for one subject\n'
                '5.Enter using ap \n6.Enter per sub per stud'), f'{class_name} {year} Term marks entry')

        if opt == 1:
            enter_marks(students_class, register, subjects, overall)

        elif opt == 2:
            sub_mark(students_class, register, subjects, overall)

        elif opt == 3:
            edit_mark(students_class, subjects, overall)

        elif opt == 4:
            sub_marks(students_class, register, subjects, overall)

        elif opt == 5:
            # info.log += '\n' +str(d.date.today()) +' - ' + str(register, students_class)
            # ap(students_class, register, subjects, overall)
            pass
        elif opt == 6:
            sub_marks3(students_class, register, subjects, overall)
        elif opt == 0:
            return data
        compile_stud1(data, stream_name, subjects, term, expected_pass, year)


def record_mark_xl(students_class, register, stud, overall):  # stud dict of names and overalls
    # info.log += '\n' +str(d.date.today()) +' - ' + str('this is reg', register)
    for name in register:
        if name not in students_class:
            students_class[name] = StudentTermMark(register[name], name)
        if name in stud:
            # info.log += '\n' +str(d.date.today()) +' - ' + str('tapinda', name)
            students_class[name].overall = stud[name]
            class_update(students_class[name])
            overall[name] = students_class[name].overall
    # info.log += '\n' +str(d.date.today()) +' - ' + str('done!!!')
    return students_class, overall


def sheet_layout(sheet1, sheet2, wb, clss, book_name):  # designing sheet
    sheet1.merge_cells(start_row=1, end_row=1, start_column=1, end_column=sheet1.max_column)
    for i in range(2, sheet1.max_column + 1):
        # sheet.merge_cells(start_row=1, end_row=2, start_column=i, end_column=i)
        sheet1.cell(2, i).alignment = Alignment(wrap_text=False, text_rotation=45)
        sheet1.column_dimensions[xl.utils.get_column_letter(i)].width = 3
    sheet1.column_dimensions['A'].width = 21
    syd = Side(style="thin")
    b_s = Border(left=syd, right=syd, top=syd, bottom=syd)

    for i in range(2, sheet1.max_row + 1):
        sheet1.row_dimensions[i].height = 10
        for j in range(1, sheet1.max_column + 1):
            sheet1.cell(i, j).border = b_s
            if sheet1.cell(i, j).font != Font(color='FF0000', size=8,name='Times New Roman'):
                sheet1.cell(i, j).font = Font(size=8,name='Times New Roman')
    sheet1.row_dimensions[2].height = 30

    btm = dsp_bottom(clss)
    tp = dsp_top(clss)
    num = 2
    row = 0
    raw_line = btm.split('\n')
    if len(raw_line) > 0:
        for item in raw_line:
            row += 1
            column = num
            column_line = item.split('\t')
            if len(column_line) > 0:
                for i in column_line:
                    column += 1
                    sheet2.cell(row, column).value = i
    num = 0
    row = 0
    raw_line = tp.split('\n')
    if len(raw_line) > 0:
        for item in raw_line:
            row += 1
            column = num
            column_line = item.split('\t')
            if len(column_line) > 0:
                for i in column_line:
                    column += 1
                    sheet2.cell(row, column).value = i
    sheet2.merge_cells(start_row=1, end_row=1, start_column=1, end_column=2)
    sheet2.merge_cells(start_row=1, end_row=1, start_column=3, end_column=5)
    sheet2.column_dimensions['A'].width = 3
    sheet2.column_dimensions['B'].width = 30
    sheet2.column_dimensions['C'].width = 3
    sheet2.column_dimensions['D'].width = 30
    sheet2.column_dimensions['E'].width = 3
    syd = Side(style="thin")
    b_s = Border(left=syd, right=syd, top=syd, bottom=syd)
    for i in range(1, sheet2.max_row + 1):
        for j in range(1, sheet2.max_column + 1):
            sheet2.cell(i, j).border = b_s
    wb.save(str(book_name))


def class_book1(clss, stream, year, term, subjects):
    info.log += '\n' + str(d.date.today()) + ' - ' + str(stream.name)
    r = 0
    c = 1
    head = stream.name + '_' + str(year) + 'Term' + str(term) + 'overall_mark_schedule'
    bk_name = head + '.xlsx'
    sys_path = os.path.dirname(os.path.dirname(__file__))
    bk_name = os.path.join(sys_path, 'files', str(year), stream.name, bk_name)
    wb1 = xl.Workbook()
    sht1 = wb1.create_sheet('Sheet1')
    sht2 = wb1.create_sheet('Sheet2')
    del wb1['Sheet']
    r += 1
    sht1.cell(r, c).value = head
    r += 1
    sht1.cell(r, c).value = stream.pass_rate
    sht1.cell(r + 1, c).value = "stream subject average"
    sht1.cell(r + 2, c).value = 'class subject average'
    for s in subjects:
        c += 1
        sht1.cell(r, c).value = s
        sht1.cell(r + 1, c).value = stream.sub_av.get(s, ' ')
        sht1.cell(r + 2, c).value = clss.sub_av.get(s, ' ')
    sht1.cell(r, c + 1).value = 'overall mark'
    sht1.cell(r, c + 2).value = "average mark"
    sht1.cell(r, c + 3).value = 'subjects passed'
    sht1.cell(r, c + 5).value = 'overall position'
    sht1.cell(r, c + 4).value = 'class position'
    r = 4
    for class_name in stream.class_list:
        ro = 0
        co = 1
        clss = stream.class_list[class_name]
        students = clss.students_list
        head = class_name + '_' + str(year) + 'Term' + str(term) + 'mark_schedule'
        book_name = head + '.xlsx'
        sys_path = os.path.dirname(os.path.dirname(__file__))
        book_name = os.path.join(sys_path, 'files', str(year), stream.name, class_name, book_name)
        wb = xl.Workbook()
        sheet1 = wb.create_sheet('Sheet1')
        sheet2 = wb.create_sheet('Sheet2')
        del wb['Sheet']
        ro += 1
        sheet1.cell(ro, co).value = head
        ro += 1
        sheet1.cell(ro, co).value = clss.pass_rate
        sheet1.cell(ro + 1, co).value = "stream subject average"
        sheet1.cell(ro + 2, co).value = 'class subject average'
        for s in subjects:
            co += 1
            sheet1.cell(ro, co).value = s
            sheet1.cell(ro + 1, co).value = stream.sub_av.get(s, ' ')
            sheet1.cell(ro + 2, co).value = clss.sub_av.get(s, ' ')

        sheet1.cell(ro, co + 1).value = 'overall mark'
        sheet1.cell(ro, co + 2).value = "average mark"
        sheet1.cell(ro, co + 3).value = 'subjects passed'
        sheet1.cell(ro, co + 5).value = 'overall position'
        sheet1.cell(ro, co + 4).value = 'class position'
        ro = 4

        for stud in students:
            c = 1
            r += 1
            co = 1
            ro += 1
            stud_name = students[stud].name
            sheet1.cell(ro, co).value = stud_name
            sht1.cell(r, c).value = stud_name
            subs = clss.students_list[stud].overall
            for sub in subjects:
                co += 1
                c += 1
                sheet1.cell(ro, co).value = subs.get(sub, ' ')
                sht1.cell(r, c).value = subs.get(sub, ' ')
                if subs.get(sub, ' ') != '':
                    if int(subs.get(sub, 0)) < 50:
                        sheet1.cell(ro, co).font = Font(color='FF0000', size=8,name='Times New Roman')
                        sht1.cell(r, c).font = Font(color='FF0000', size=8, name='Times New Roman')

                    else:
                        sheet1.cell(ro, co).font = Font(size=8, name='Times New Roman')
                        sht1.cell(r, c).font = Font(size=8, name='Times New Roman')

            sheet1.cell(ro, co + 5).value = stream.overall_position.get(stud_name, ' ')
            sheet1.cell(ro, co + 4).value = clss.class_position.get(stud_name, ' ')
            sheet1.cell(ro, co + 1).value = students[stud].overall_mark
            sheet1.cell(ro, co + 2).value = clss.average_marks.get(stud_name, ' ')
            sheet1.cell(ro, co + 3).value = students[stud].passed

            sht1.cell(r, c + 5).value = clss.overall_position.get(stud_name, ' ')
            sht1.cell(r, c + 4).value = clss.class_position.get(stud_name, ' ')
            sht1.cell(r, c + 1).value = students[stud].overall_mark
            sht1.cell(r, c + 2).value = clss.average_marks.get(stud_name, ' ')
            sht1.cell(r, c + 3).value = students[stud].passed
        md(__file__, f'files/{year}/{stream.name}/{class_name}')
        sheet_layout(sheet1, sheet2, wb, clss, book_name)
    sheet_layout(sht1, sht2, wb1, stream, bk_name)


def run2(year, stream_name, term, class_name, data):
    opt = 99
    stream = data[year].stream_list[stream_name].term_marks.get(term)
    # info.log += '\n' +str(d.date.today()) +' - ' + str(stream)
    # info.log += '\n' +str(d.date.today()) +' - ' + str(stream.class_list)
    class_list = stream.class_list
    class_info = class_list[class_name]
    if stream is None:
        info.log += '\n' + str(d.date.today()) + ' - ' + str('hoyoo zvadhakwa')
    while opt != 60:
        # info.log += '\n' +str(d.date.today()) +' - ' + str(class_name,year)
        subjects = data[year].stream_list[stream_name].subjects
        expected_pass = data[year].stream_list[stream_name].expected_pass
        opt = app_opt(prompt_list(f'1. Enter marks\n'
                                  '2. display info\n'
                                  '3. extract subjects'
                                  ), f'{class_name} {year} Terms marks')
        students_class = class_info.students_list  #
        overall = class_info.overall_marks  #
        register = data[year].stream_list[stream_name].class_list[class_name].class_register
        if register != {}:
            register = register.registered_pupils
        else:
            data[year].stream_list[stream_name].class_list[class_name].class_register = RegisterProfile(class_name)
            register = data[year].stream_list[stream_name].class_list[class_name].class_register.registered_pupils
        if opt == 0:
            break

        elif opt == 1:
            data = mark_opt(data, students_class, stream_name, class_name, year, term, register, subjects,
                            overall, expected_pass)
        elif opt == 2:
            display_opt(data, students_class, stream, class_info, stream_name, class_name, year, term, subjects)

        elif opt == 3:
            stud = ext_xl(subjects, data, stream_name, class_name, year, term)
            record_mark_xl(students_class, register, stud, overall)
            info.log += '\n' + str(d.date.today()) + ' - ' + str('oh yes')
            compile_stud1(data, stream_name, subjects, term, expected_pass, year)
            pass

        class_book1(class_info, stream, year, term, subjects)
        save_data(data, 'data_year', 0)


# ex and test marks
# create record book object
class Exercise:
    def __init__(self, name, date, out_of, typ):
        self.typ = typ
        self.name = name
        self.date = date
        self.out_of = out_of
        self.lowest_mark = ''
        self.highest_mark = ''
        self.f_mark = ''
        self.marks = {}
        #self.p_marks = {}
        self.average_mark = 0
        #self.average_p_mark = 0
        self.passed = {}
        self.failed = {}
        self.highest = {}
        self.lowest = {}
        self.missing_marks = []
        #self.top_ten = {}
        #self.bottom = {}

def edit_mark1(register, date, of, marks, state=False):
    while not state:
        name = app_txt(register)
        if name in ['exit', None, '0', 0]:
            break
        info.log += '\n' + str(d.date.today()) + ' - ' + str(f'Entering marks for \n{name}')
        if date == 0 or date is None:
            break
        mark = marks.get(name, 'xx')
        while True:
            mark = get_int('mark', f'current mark for {name} is {mark} \nfor exercise on {date}')
            if mark == 'e':
                return
            elif str(mark).isdigit():
                mark = int(mark)
                if mark <= of:
                    marks[name] = mark
                    info.log += '\n' + str(d.date.today()) + ' - ' + str('the mark was entered')
                break


def sub_marks1(record: dict, register: dict, date: str):
    """
    :param record: dictionary with ex objects
    :param register: dictionary with students and students classes
    :param date: date of the exercise
    :return: excise object
    """

    marks = record[date].marks
    of = record[date].out_of
    dts = {name: marks[name] for name in marks}
    mark_for_studs = ap2([x for x in register], dts, 0)
    # info.log += '\n' +str(d.date.today()) +' - ' + str(mark_for_studs)
    for name in mark_for_studs:
        mark1 = mark_for_studs[name]
        if str.isdecimal(str(mark1)):
            mark = int(mark1)
            if mark > of:
                info.log += '\n' + str(d.date.today()) + ' - ' + str('error value surpasses the overall')
                continue
            else:
                marks[name] = mark
        else:
            continue
    info.log += '\n' + str(d.date.today()) + ' - ' + str('tabuda')
    return record[date]


def book_entry(record, sheet, register, class_name, year, f_s=11):
    from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
    ro = 5
    info.log += '\n' + str(d.date.today()) + ' - ' + str(f'{sheet.max_row} {sheet.max_column}')
    # entering titles for the rows
    if sheet.max_row == 1 and sheet.max_column == 1:
        sheet.cell(1, 1).value = class_name + str(year)
        sheet.cell(2, 1).value = 'name'
        sheet.cell(3, 1).value = 'date'
        sheet.cell(4, 1).value = 'out_of'
        info.log += '\n' + str(d.date.today()) + ' - ' + str('done')
    for name in register:
        if sheet.cell(ro, 1).value is None:
            sheet.cell(ro, 1).value = name
        ro += 1
    sheet.cell(ro, 1).value = 'highest_mark'
    sheet.cell(ro + 1, 1).value = 'lowest_mark'
    sheet.cell(ro + 2, 1).value = 'average_mark'
    sheet.cell(ro + 3, 1).value = 'frequent_mark'
    sheet.cell(ro + 4, 1).value = 'pass_rate'
    for date in sorted(record):
        ro = 2
        co = sheet.max_column + 1
        if sheet.cell(ro, 1).value == 'name':
            sheet.cell(ro, co).value = record[date].name
            ro += 1
        if sheet.cell(ro, 1).value == 'date':
            sheet.cell(ro, co).value = date
            ro += 1
        if sheet.cell(ro, 1).value == 'out_of':
            sheet.cell(ro, co).value = record[date].out_of
            ro += 1
        for name in register:
            if sheet.cell(ro, 1).value == name:
                info.log += '\n' + str(d.date.today()) + ' - ' + str('apa tiri kusvika')
                if sheet.cell(ro, co).value is None:
                    if record[date].typ == 'Exercise' or record[date].typ == 'E' :
                        sheet.cell(ro, co).value = record[date].marks.get(name, '')
                        info.log += '\n' + str(d.date.today()) + ' - ' + str(
                            'this is the ' + str(record[date].marks.get(name, '')))
                        if record[date].marks.get(name, '') != '':
                            if int(record[date].marks.get(name)) < int(record[date].out_of) / 2:
                                sheet.cell(ro, co).font = Font(color='FF0000', size=f_s,name='Times New Roman')
                                sheet.cell(ro, co).fill = PatternFill(start_color='FFC5C5', end_color="FFC5C5",
                                                                      fill_type='solid')

                            else:
                                sheet.cell(ro, co).font = Font(size=f_s,name='Times New Roman')
                    elif record[date].typ == 'Test' or record[date].typ == 'T':
                        sheet.cell(ro, co).value = record[date].p_marks.get(name, '')
                        if record[date].p_marks.get(name) != '':
                            if int(record[date].marks.get(name, 0)) < int(record[date].out_of) / 2:
                                sheet.cell(ro, co).font = Font(color='FF0000', size=f_s,name='Times New Roman')
                                sheet.cell(ro, co).fill = PatternFill(start_color = 'FFC5C5', end_color = "FFC5C5", fill_type = 'solid')
                                # info.log += '\n' +str(d.date.today()) +' - ' + str('red mark')
                            else:
                                sheet.cell(ro, co).font = Font(size=f_s,name='Times New Roman')
            ro += 1
        sheet.cell(ro, co).value = record[date].highest_mark
        sheet.cell(ro + 1, co).value = record[date].lowest_mark
        sheet.cell(ro + 2, co).value = record[date].average_mark
        sheet.cell(ro + 3, co).value = record[date].f_mark
        #sheet.cell(ro + 4, co).value = record[date].pass_rate

        # designing the sheet

    sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=sheet.max_column)
    lines = 16 - sheet.max_column
    if lines >= 0 :
        max_column = sheet.max_column + lines
    else:
        max_column = sheet.max_column
    for i in range(2,max_column + 1):
        # sheet.merge_cells(start_row=1, end_row=2, start_column=i, end_column=i)
        sheet.cell(2, i).alignment = Alignment(wrap_text=False, text_rotation=45)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 4
    sheet.column_dimensions['A'].width = 30
    syd = Side(style="thin")
    b_s = Border(left=syd, right=syd, top=syd, bottom=syd)

    for i in range(2, sheet.max_row + 1):
        sheet.row_dimensions[i].height = 14.4
        for j in range(1, sheet.max_column + 1):
            sheet.cell(i, j).border = b_s
            if sheet.cell(i, j).font != Font(color='FF0000', size=11,name='Times New Roman'):
                if i == 3:
                    sheet.cell(i, j).font = Font(size=9,name='Times New Roman')
                else:
                    sheet.cell(i, j).font = Font(size=11,name='Times New Roman')
    sheet.row_dimensions[2].height = 60


def enter_multi_dates(record, register):
    """
        :param record: dictionary with ex objects
        :param register: dictionary with students and students classes
        :return: excise object
        """
    for name in register:
        for date in record:
            of = record[date].out_of
            mark = record[date].marks.get(name, 'xx')
            if mark == 'xx':
                while True:
                    mark = get_int('mark',
                                   f'current mark for {name} is {mark} \nfor exercise on {date} \n{record[date].name}')
                    if mark == 'e':
                        return
                    elif str(mark).isdigit():
                        mark = int(mark)
                        if mark <= of:
                            marks = record[date].marks
                            marks[name] = mark
                            break
                    else:
                        break
            update_ex(record[date])
            record[date].missing_marks = [name for name in register if name not in record[date].marks]


def enter_marks1(marks, register, of, sub_name, edit=False):
    mark1 = -1
    mark = -1
    for name in register:
        while True:
            if not edit:
                m = marks.get(name, "xx")
                if m != 'xx':
                    break
            mark1 = get_int('mark', f'{sub_name} for a {name} \n  current mark {marks.get(name, "xx")}')
            if str.isdecimal(str(mark1)):
                mark = int(mark1)
                if mark > of:
                    info.log += '\n' + str(d.date.today()) + ' - ' + str('error value surpasses the overall')
                    continue
                break
            else:
                mark = -1
                info.log += '\n' + str(d.date.today()) + ' - ' + str('\nM0001 error\n')
                if mark1 == 'e':
                    break
                elif mark1 == 's':
                    break
                continue
        if mark == -1:
            if mark1 == 'e':
                break
            elif mark1 == 's':
                continue
            continue
        if marks.get(name) is None:
            marks[name] = mark  # student classes


def avg_ex(marks):
    if len(marks) == 0:
        return 0
    avg_mark = 0
    for name in marks:
        mark = marks[name]
        avg_mark += int(mark)
    avg_mark /= len(marks)
    return avg_mark


def grade_ex(marks, of):
    passed, failed = {}, {}
    for name in marks:
        mark = marks[name]
        mark = int(mark)
        if mark >= of / 2:
            passed[name] = mark
        elif mark < of:
            failed[name] = mark
    return passed, failed


def p(marks):
    high = {}
    low = {}
    for name in marks:
        mark = marks[name]
        high[name] = mark
        low[name] = -mark
    high = class_position(high)
    low = class_position(low)
    for name in marks:
        if high[name] == 1:
            high[name] = marks[
                name]
        else:
            del high[name]
        if low[name] == 1:
            low[name] = marks[name]
        else:
            del low[name]
    return high, low


def freq_mark(marks):
    freq = {x: marks.count(x) for x in marks}
    freq = p(freq)[0]
    info.log += '\n' + str(d.date.today()) + ' - ' + str(freq)
    a = 0
    for x in freq.keys():
        a += x
    if len(freq) > 0:
        a /= len(freq)
    mark = round(a)
    return mark


def update_ex(ex):
    ex.p_marks= {name: round((ex.marks[name] * 100) / ex.out_of) for name in ex.marks}
    ex.average_p_mark = round(avg_ex(ex.p_marks))
    ex.average_mark = round(avg_ex(ex.marks))
    ex.passed, ex.failed = grade_ex(ex.marks, ex.out_of)
    ex.highest, ex.lowest = p(ex.marks)
    if len(ex.lowest) >0:
     ex.lowest_mark = [x for x in ex.lowest.values()][0]
    else:
        ex.lowest_mark = 'no_one'
    if len(ex.highest)> 0:
        ex.highest_mark = [x for x in ex.highest.values()][0]
    else:
        ex.highest_mark = 'no_one'
    ex.f_mark = freq_mark([x for x in ex.marks.values()])
    if len(ex.marks)> 0:
        ex.pass_rate = round(len(ex.passed)*100/len(ex.marks))
    else :
        ex.pass_rate = 0



class Sub:
    def __init__(self, name):
        self.name = name
        self.ex_record = {}
        self.test_record = {}
        self.term_mark = {}


def run3(year, stream_name, term, class_name, data):
    stream = data[year].stream_list[stream_name].term_marks[term]
    class_list = stream.class_list
    class_info = class_list[class_name]

    while True and (stream_name != 0 and stream_name is not None and class_name != 0):
        info.log += '\n' + str(d.date.today()) + ' - ' + str(stream_name)
        ex = 0
        register = data[year].stream_list[stream_name].class_list[class_name].class_register
        if register == {}:
            break

        register = register.registered_pupils
        if len(register) == 0:
            break

        subjects = data[year].stream_list[stream_name].subjects
        subs = ['add subject']
        subs.extend([x for x in class_info.subjects])
        sub_name = app_txt(subs)
        if sub_name in [0, None, 'e', 's', '0']:
            break
        elif sub_name == 'add subject':
            sub_name = app_txt([x for x in subjects if x not in class_info.subjects])
            if sub_name in [0, None, 'e', 's', '0']:
                break
            info.log += '\n' + str(d.date.today()) + ' - ' + str(class_info.subjects)
            class_info.subjects[sub_name] = Sub(sub_name)
            continue
        else:
            test_record = class_info.subjects[sub_name].test_record
            ex_record = class_info.subjects[sub_name].ex_record
            while True:
                typ = app_txt(['Test', 'Exercise'])
                if typ is None:
                    break
                new_date = 0
                while True:
                    opt = app_opt(prompt_list('1.Enter all marks using entered date\n2.enter new date \n'
                                              "3. Enter for multiple dates \n4.Edit ex_info"
                                              ), f'{class_name} {sub_name} term{term[0]} {year} \n{typ} marks entry')
                    if opt == 0:
                        break

                    if opt == 1:
                        while True:
                            date = None
                            if typ == 'Test':
                                date = app_txt(test_record)
                                if date in [0, '0', None, 'e']:
                                    break
                                ex = test_record[date]
                            elif typ == 'Exercise':

                                date = app_txt(ex_record)
                                if date in [0, '0', None, 'e']:
                                    break
                                ex = ex_record[date]
                            info.log += '\n' + str(d.date.today()) + ' - ' + str('tapinda')
                            if date in [0, '0', None, 'e']:
                                break
                            while True:
                                opt = app_opt(
                                    prompt_list(
                                        '1.Enter marks per student \n2.Edit marks '
                                        '\n3.print missing marks \n4.Enter all marks'
                                    ), f'{sub_name}\n {ex.date}{year}-{ex.name} \n{typ} marks entry')
                                if opt == 1:
                                    enter_marks1(ex.marks, register, ex.out_of, sub_name)
                                elif opt == 2:
                                    edit_mark1(register, date, ex.out_of, ex.marks)
                                elif opt == 3:
                                    app_display({name: 'no mark' for name in ex.missing_marks})
                                elif opt == 4:
                                    if typ == 'Test':
                                        ex = sub_marks1(test_record, register, date)
                                        if ex is None:
                                            break
                                    elif typ == 'Exercise':
                                        ex = sub_marks1(ex_record, register, date)
                                        if ex is None:
                                            break
                                    else:
                                        break
                                elif opt == 0:
                                    break
                                book_save(ex_record, test_record, year, class_name, term, sub_name, register)
                                save_data(data, 'data_year', 0)

                    elif opt == 2:
                        date = app_input1('date', f'enter {typ} date (mmdd)')
                        if date in [0, '0', None, 'e']:
                            break
                        if len(date) == 4 and str.isdigit(date):
                            if typ == "Exercise":
                                if ex_record.get(date) is None:
                                    new_date = 1
                                else:
                                    ex = ex_record[date]
                                    info.log += '\n' + str(d.date.today()) + ' - ' + str(f'aya {ex_record[date].marks}')

                            elif typ == 'Test':
                                if test_record.get(date) is None:
                                    new_date = 1
                                else:
                                    ex = test_record[date]

                        else:
                            info.log += '\n' + str(d.date.today()) + ' - ' + str('enter correct date')
                            continue
                        if new_date == 1:
                            b = 0
                            while True:
                                if b == 1:
                                    break
                                topic_name = app_input1('name', 'enter topic name')
                                topic_name = str(topic_name)
                                if str.isdigit(topic_name):
                                    info.log += '\n' + str(d.date.today()) + ' - ' + str('write a proper name')
                                    break
                                while True:
                                    of = get_int('mark', 'enter out of \n')
                                    if of in [0, '0', None, 'e']:
                                        b = 1
                                        break
                                    of = int(of)
                                    if str.isdigit(str(of)):
                                        ex = Exercise(topic_name, date, of, typ)
                                        if typ == 'Exercise':
                                            ex_record[ex.date] = ex

                                        elif typ == 'Test':
                                            test_record[ex.date] = ex
                                        b = 1
                                        break

                        if ex != 0:
                            info.log += '\n' + str(d.date.today()) + ' - ' + str('we did it')
                            marks = ex.marks
                            ex.missing_marks = [name for name in register if name not in ex.marks]
                            info.log += '\n' + str(d.date.today()) + ' - ' + str(marks)
                            of = ex.out_of
                            opt = app_opt(
                                prompt_list(
                                    '1.Enter all marks for all students and all subjects \n2.Edit marks for student '
                                    'per subject'"\n3.Print missing exercise"), f'{ex.name} {ex.date} marks entry')

                            if opt == 1:
                                enter_marks1(marks, register, of, sub_name)
                            elif opt == 2:
                                edit_mark1(register, date, of, marks)
                            elif opt == 3:
                                app_display({name: 'no mark' for name in ex.missing_marks})
                            if ex.marks != {}:
                                update_ex(ex)
                                ex.missing_marks = [name for name in register if name not in ex.marks]

                            book_save(ex_record, test_record, year, class_name, term, sub_name, register)
                            save_data(data, 'data_year', 0)

                    elif opt == 3:
                        if typ == 'Test':
                            enter_multi_dates(test_record, register)
                            book_save(ex_record, test_record, year, class_name, term, sub_name, register)
                            save_data(data, 'data_year', 0)
                            if ex is None:
                                break
                        elif typ == 'Exercise':
                            enter_multi_dates(ex_record, register)
                            book_save(ex_record, test_record, year, class_name, term, sub_name, register)
                            save_data(data, 'data_year', 0)
                            if ex is None:
                                break
                        else:
                            break
                    elif opt == 4:
                        while True:
                            date = None
                            record = {}
                            if typ == 'Test':
                                date = app_txt(test_record)
                                if date in [0, '0', None, 'e']:
                                    break
                                ex = test_record[date]
                                record = test_record
                            elif typ == 'Exercise':
                                date = app_txt(ex_record)
                                if date in [0, '0', None, 'e']:
                                    break
                                ex = ex_record[date]
                                record = ex_record
                            info.log += '\n' + str(d.date.today()) + ' - ' + str('tapinda')
                            if date in [0, '0', None, 'e']:
                                break
                            if date in record:
                                new_date = app_input1('date', f'enter {typ} new date (mmdd)')
                                if new_date in [0, '0', None, 'e']:
                                    break

                                elif len(new_date) == 4 and str.isdigit(date):
                                    if new_date not in record:
                                        record[new_date] = record.pop(date)
                                        record[new_date].date = new_date
                                        #print('edited')
                                    else:
                                        #print('date already exist')
                                        pass
                                else:
                                    #print('incorrect date format')
                                    pass

                        pass


def book_save(ex_record, test_record, year, class_name, term, sub_name, register):
    s_ex_record = dict(sorted(ex_record.items()))
    s_test_record = dict(sorted(test_record.items()))
    ex_record.clear()
    test_record.clear()
    ex_record.update(s_ex_record)
    test_record.update(s_test_record)
    del s_test_record
    del s_ex_record
    book_name = class_name + str(year) + term + sub_name + 'mark_record_book' + '.xlsx'
    md(__file__, 'record_of_marks')
    sys_path = os.path.dirname(os.path.dirname(__file__))
    path = os.path.join(sys_path, 'record_of_marks')
    book_name = os.path.join(path, book_name)
    wb = xl.Workbook()
    test_sheet = wb.create_sheet('Tests')
    ex_sheet = wb.create_sheet('Exercises')
    #ex_sheet = wb.create_sheet('Exercises')
    del wb['Sheet']
    info.log += '\n' + str(d.date.today()) + ' - ' + str('entering the marks in the book')
    roots = {}
    root_num = -1
    for i in ex_record:
        root_num += 1
        count = root_num // 15
        if len(roots) == count:
            roots[count] = {}
        roots[count][i]= ex_record[i]
    for i in roots:
        ex_sheets = wb.create_sheet('Exercises' + str(i))
        record = roots[i]
        book_entry(record, ex_sheets, register, class_name, year)
    book_entry(ex_record, ex_sheet, register, class_name, year)
    book_entry(test_record, test_sheet, register, class_name, year)
    wb.save(book_name)
    wb.close()


# display opt
def dsp_bottom1(clss):
    top_ten = clss.bottom_ten
    app_display(top_ten)


def dsp_top1(clss):
    top_ten = clss.top_ten
    app_display(top_ten)


def dsp_top(clss):
    head = "top ten \n"
    top_ten = clss.top_ten
    for stud in top_ten:
        head += str(top_ten[stud]) + '\t' + str(stud) + "\n"
    return head


def t_sheet(head, book_name):
    wb = xl.load_workbook(book_name)
    sheet = wb['Sheet2']
    sheet = cell_entry(wb, book_name, head, sheet, 0)
    return sheet


def b_sheet(head, book_name, num):
    wb = xl.load_workbook(book_name)
    sheet = wb['Sheet2']
    sheet = cell_entry(wb, book_name, head, sheet, num)
    return sheet


def symbol(num, level):
    if 2 < level < 5:
        num = int(num)
        if 70 <= num <= 100:
            return 'A'
        elif 60 <= num <= 69:
            return 'B'
        elif 50 <= num <= 59:
            return 'C'
        elif 45 <= num <= 49:
            return 'D'
        elif 40 <= num <= 44:
            return 'E'
        elif 0 <= num <= 39:
            return 'U'
        else:
            return " "
    elif 0 < level < 3:
        if 75 <= num <= 100:
            return 1
        elif 70 <= num <= 74:
            return 2
        elif 65 <= num <= 70:
            return 3
        elif 60 <= num <= 64:
            return 4
        elif 55 <= num <= 59:
            return 5
        elif 50 <= num <= 54:
            return 6
        elif 45 <= num <= 49:
            return 7
        elif 40 <= num <= 44:
            return 8
        elif 0 <= num <= 39:
            return 9
        else:
            return " "
    elif 4 < level < 7:
        if 80 <= num <= 100:
            return 5
        elif 70 <= num <= 79:
            return 4
        elif 60 <= num <= 69:
            return 3
        elif 50 <= num <= 59:
            return 2
        elif 45 <= num <= 49:
            return 1
        elif 35 <= num <= 44:
            return 0
        elif 0 <= num <= 34:
            return 'F'
        else:
            return " "


def display_stream(stream, year, term, subjects):
    stream_sub_average = '0'
    t = '\t'
    n = '\n'
    heading = '\n'
    heading2 = '\n'
    class_pass_rate = stream.pass_rate
    class_list = stream.class_list
    heading2 += stream.name + str(year) + str(term) + t + str(class_pass_rate) + n
    c = 0
    for cl in class_list:
        clss = class_list[cl]
        class_name = clss.name
        students = clss.students_list
        heading += class_name + str(year) + str(term) + t + str(class_pass_rate) + n
        stream_sub_average = ""
        class_sub_average = ""

        for s in subjects:
            sam = stream.sub_av.get(s, ' ')
            cam = clss.sub_av.get(s, ' ')
            heading += s[:3] + t
            if c == 0:
                heading2 += s[:3] + t

            stream_sub_average += str(sam) + t
            class_sub_average += str(cam) + t

        heading += 'ov' + t + 'cl' + t + 'om' + t + 'av' + n
        if c == 0:
            heading2 += 'ov' + t + 'cl' + t + 'om' + t + 'av' + n
        c = 1
        for stud in students:
            stud_name = students[stud].name
            ov_pos = stream.overall_position.get(stud_name, ' ')
            cl_pos = clss.class_position.get(stud_name, ' ')
            ov_mark = clss.students_list[stud].overall_mark
            av_mark = clss.average_marks.get(stud_name, ' ')
            subs = clss.students_list[stud].overall
            marks = ''
            for sub in subjects:
                mark = subs.get(sub, ' ')
                marks += str(mark) + t
            marks += str(ov_pos) + t + str(cl_pos) + t + str(ov_mark) + t + str(av_mark) + t + stud_name + n
            heading += marks
            heading2 += marks
        heading += '\n' + class_sub_average + ' ' + t + ' ' + t + ' ' + t + ' ' + t + "class_average mark" + n
    heading += stream_sub_average + ' ' + t + ' ' + t + ' ' + t + ' ' + t + 'stream average mark' + n
    heading2 += stream_sub_average + ' ' + t + ' ' + t + ' ' + t + ' ' + t + 'stream average mark' + n
    return heading, heading2


def get_reports(stream, clss, year, term, subjects, class_info):
    text = ''
    for stud in clss.students_list:
        text += get_report(stream, clss, stud, year, term, subjects, class_info)
    r_name = clss.name + '_' + stream.name + '_' + '_' + str(year) + '_' + 'reports' + '.txt'
    sys_path = os.path.dirname(os.path.dirname(__file__))
    md(__file__, f'files/{str(year)}/{stream.name}/{clss.name}/Reports')
    md(__file__, f'files/{str(year)}/{stream.name}/{clss.name}/Reports/' + 'Term' + str(term))
    name = os.path.join(sys_path,
                        f'files/{str(year)}/{stream.name}/{clss.name}/Reports/' + 'Term' + str(term) + '/' + r_name)
    r = open(name, 'w')
    r.write(text)
    r.close()
    return text


def get_report(stream, clss, stud_name, year, term, saba, class_info):
    form = clss.name
    ov = stream.overall_position
    om = clss.students_list[stud_name].overall_mark
    am = clss.average_marks
    att = class_info.class_register.att_register['term_' + term[0]]
    if att != {}:
        att = att.students_totals

    results = clss.students_list[stud_name].overall
    sub_av = clss.sub_av
    t = '\t'
    n = '\n'
    text = n + '{:*^62}'.format(stud_name) + n

    text += stud_name + n + 'FORM' + t * 2 + str(form) + t * 2 + 'TERM' + t + str(term[0]) + t + 'YEAR' + t + str(
        year) + n
    text += 'NO IN FORM' + t + str(len(ov)) + t * 2 + 'POSITION IN FORM' + t + str(ov.get(stud_name, 'xx')) + n
    text += 'OVERALL MARK' + t + str(om) + t * 2 + 'AVERAGE MARK' + t + str(am.get(stud_name, 'xx')) + n
    text += 'ATTENDANCE' + t + str((att.get(stud_name, {})).get('p', 0)) + t * 2 + 'OUT OF' + t * 2 + str(
        (att.get(stud_name, {})).get('p', 0) + (att.get(stud_name, {})).get('a', 0)) + t + "DAYS"
    text += n
    text += n
    text += 'SUBJECTS' + " " * (
            12 - len('SUBJECTS')) + t + 'EX/MARK' + t + 'AVERAGE' + t + 'SYMBOL' + t + "TEACHER'S COMMENT" + n
    for sub in saba:
        if sub in results:
            if len(sub) < 16:
                name = sub + " " * (16 - len(sub))
            else:
                name = sub
            sub_name = name[:15] + t + str(results.get(sub, 'xx')) + t + str(sub_av.get(sub, 'xx')) + t + str(
                symbol(results[sub], int(form[:1]))) + t + '_' * 22 + n
            text += sub_name
    text += n
    text += "Teachers' remarks and signature" + '.' * 30 + n
    text += "Headmaster's Comment" + '.' * 42 + n
    text += "Parent/Guardian's signature" + '.' * 35 + n
    text += n * 3 + 'STAMP'
    text += n * 2 + '-' * 65
    r_name = stud_name + '_' + form + '_' + str(year) + '.txt'
    sys_path = os.path.dirname(os.path.dirname(__file__))
    md(__file__, f'files/{str(year)}/{stream.name}/{form}/Reports')
    md(__file__, f'files/{str(year)}/{stream.name}/{form}/Reports/' + 'Term' + str(term))
    name = os.path.join(sys_path,
                        f'files/{str(year)}/{stream.name}/{form}/Reports/' + 'Term' + str(term) + '/' + r_name)
    r = open(name, 'w')
    r.write(text)
    r.close()
    return text


def display_opt(data, students_class, stream, clss, stream_name, class_name, year, term, subjects):
    class_info = data[year].stream_list[stream_name].class_list[class_name]
    get_reports(stream, clss, year, term, subjects, class_info)
    opt = app_opt(prompt_list('1.stream statistics\n'
                              '2.position\n'
                              '3.top ten\n'
                              '4.bottom ten \n'
                              '5.marks\n'), f'{class_name}{year} marks display')
    if opt == 1:
        heading, heading1 = display_stream(stream, year, term, subjects)
        # info.log += '\n' +str(d.date.today()) +' - ' + str(heading)
        cell_entry1(heading1, 0)
        book_name = stream.name + '_' + str(year) + '_' + str(term) + 'overall_states.xlsx'
        sys_path = os.path.dirname(os.path.dirname(__file__))
        book_name = os.path.join(sys_path, 'files', str(year), stream.name, book_name)
        wb = xl.Workbook()
        wb.create_sheet('Sheet1')
        wb.create_sheet('Sheet2')
        del wb['Sheet']
        wb.save(str(book_name))
        marks_book(heading1, book_name)
        btm = dsp_bottom(stream)
        tp = dsp_top(stream)
        b_sheet(btm, book_name, 4)
        t_sheet(tp, book_name)
        info.log += '\n' + str(d.date.today()) + ' - ' + str('opt {opt}')

    elif opt == 2:
        opt = app_opt(prompt_list('1.overall  \n2.class  \n'), "position")
        if opt == 1:
            app_display(stream.overall_position)

        elif opt == 2:
            app_display(clss.overall_position)

    elif opt == 3:
        opt = app_opt(prompt_list('1.overall \n2.class\n'), 'top ten')
        if opt == 1:
            info.log += '\n' + str(d.date.today()) + ' - ' + str(dsp_top1(stream))
        elif opt == 2:
            info.log += '\n' + str(d.date.today()) + ' - ' + str(dsp_top1(clss))

    elif opt == 4:
        opt = app_opt(prompt_list('1.overall \n2.class\n'), ' bottom ten')
        if opt == 1:
            info.log += '\n' + str(d.date.today()) + ' - ' + str(dsp_bottom1(stream))

        elif opt == 2:
            info.log += '\n' + str(d.date.today()) + ' - ' + str(dsp_bottom1(clss))

    elif opt == 5:
        info.log += '\n' + str(d.date.today()) + ' - ' + str(dsp_marks1(clss, subjects))
    elif opt == 7:
        stud_name = app_txt(data[year].stream_list[stream_name].class_list[class_name].class_register.registered_pupils)
        r = 'no report found'
        for s in students_class:
            if stud_name == s:
                r = get_report(stream, clss, stud_name, year, term, subjects, class_info)
                break
        info.log += '\n' + str(d.date.today()) + ' - ' + str(r)
    elif opt == 6:
        a = get_reports(stream, clss, year, term, subjects, class_info)
        info.log += '\n' + str(d.date.today()) + ' - ' + str(a)


# launcher
def create_stream(data, stream_name, year, term):
    name = stream_name
    if name not in data[year].stream_list[stream_name].term_marks[term].keys():
        data[year].stream_list[stream_name].term_marks[term] = StreamTermMark(name, year)
        # info.log += '\n' +str(d.date.today()) +' - ' + str(f'{name} created\n')
    else:
        info.log += '\n' + str(d.date.today()) + ' - ' + str(f'{name} already recorded\n')
    return data


def create_class(data, stream_name, name, year, term):
    name = stream_name[-1] + '_' + name.upper()
    if stream_name not in data.get(year).stream_list.keys():
        info.log += '\n' + str(d.date.today()) + ' - ' + str(f'no such stream as {stream_name} recorded')
    else:
        data[year].stream_list[stream_name].class_list[name].term_marks[term] = ClassTermMark(name, year)
        data[year].stream_list[stream_name].term_marks[term].class_list[name] = (
            data[year].stream_list[stream_name].class_list[name].term_marks)[term]
        # info.log += '\n' +str(d.date.today()) +' - ' + str(f'{name} created {stream_name}\n')
    return data


def run():
    while True:
        year, data = get_data()
        if year is None:
            break
        while True:
            if year is None:
                break
            opt = app_opt(prompt_list('1. Enroll/class allocation\n'
                                      '2. school work'), f'{year}')
            if opt == 1:
                run1(year, data)
            if opt == 0:
                break
            if opt == 2:
                while True:
                    stream_name = get_stream(year, data)
                    info.log += '\n' + str(d.date.today()) + ' - ' + str(stream_name)
                    if stream_name is None or str(stream_name) == '0':
                        break
                    stream = data[year].stream_list.get(stream_name)
                    while True:
                        class_list1 = stream.class_list
                        info.log += '\n' + str(d.date.today()) + ' - ' + str(f'{stream_name} {year} classes')
                        class_name = app_txt(class_list1)
                        if class_name is None or class_name == 0 or class_name == '0':
                            break
                        # creates folders
                        md(__file__, 'files')
                        md(__file__, f'files/{year}/{stream_name}')
                        md(__file__, f'files/{year}/{stream_name}/{class_name}')
                        terms_mark = stream.term_marks
                        info.log += '\n' + str(d.date.today()) + ' - ' + str(terms_mark)
                        term = app_txt(terms_mark)
                        if term is None:
                            break
                        stream1 = data[year].stream_list[stream_name].term_marks.get(term)
                        info.log += '\n' + str(d.date.today()) + ' - ' + str(stream1)
                        if stream1 is None or stream1 == {}:
                            data = create_stream(data, stream_name, year, term)
                            for i in data[year].stream_list[stream_name].class_list:
                                name = i[2:]
                                data = create_class(data, stream_name, name, year, term)
                        while True:
                            opt = app_opt(prompt_list('1. Terms marks\n'
                                                      '2. Ex and test marks\n'
                                                      '3. mark register'), f'{year}')
                            info.log += '\n' + str(d.date.today()) + ' - ' + str(f'the opt is {opt}')
                            if opt == 1:
                                run2(year, stream_name, term, class_name, data)
                            if opt == 2:
                                run3(year, stream_name, term, class_name, data)
                            if opt == 3:
                                run4(year, stream_name, term, class_name, data)
                                pass
                            if opt == 0:
                                break
                save_data(data, 'data_year', 0)


run()
